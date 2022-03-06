#####################################################
#                                                   #
#          Copyright(c) 2022 KeppyNaushika          #
#                                                   #
#        This software is released under the        #
#      GNU Affero General Public License v3.0,      #
#                    see LICENSE.                   #
#                                                   #
#      The github repository of this software:      #
# https://github.com/KeppyNaushika/scoring_at_once/ #
#                                                   #
#####################################################

import tkinter
import tkinter.filedialog
import tkinter.font
import tkinter.messagebox

import PIL
import PIL.Image
import PIL.ImageDraw
import PIL.ImageFont
import PIL.ImageTk

import openpyxl
import openpyxl.drawing.image
import openpyxl.styles
import openpyxl.utils.cell
import openpyxl.worksheet.datavalidation
import openpyxl.worksheet.views

import functools
import glob
import img2pdf
import json
import os
import subprocess


def nothing_to_do(*args, **kwargs):
  tkinter.messagebox.showinfo(
    "未実装", "この機能は現在実装されておりません"
  )

# class: 子ウインドウ:
class SubWindow:
  def __init__(self, parent) -> None:
    self.parent = parent
    self.window = None  

  def this_window_close(self):
    self.window.withdraw()
    self.window = None
    self.parent.destroy()
    main()
    return "break"

  def sub_window_loop(func):
    def inner(self, *args, **kargs):
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      if dict_config["index_projects_in_listbox"] is not None:
        dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
        path_dir    : str = dict_project["path_dir"]
        if os.path.exists(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx"):
          bool_del_xlsx = tkinter.messagebox.askokcancel(
            "配点ファイルが存在しています", 
            "配点ファイルに入力した情報を保存するには, ［配点を読み込む］をクリックする必要があります. \n"
            + "既に Excel で配点を入力されている場合で［配点を読み込む］をクリックしていない場合は, 入力した情報が破棄されます. \n\n"
            + "入力した配点を保存した上で操作を続行したい場合は, ［キャンセル］をクリックした後, ［配点を読み込む］をクリックして配点を読み込んでから, もう一度実行して下さい. \n\n"
            + "配点ファイルを削除してもよろしいですか？"
          )
          if not bool_del_xlsx:
            return None
          else:
            try:
              os.remove(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx")
            except PermissionError:
              tkinter.messagebox.showerror(
                "ファイルを削除できません",
                "ファイルを削除できませんでした. \n"
                + "ファイルを開いていませんか？\n"
                + "Excel を終了して, もう一度お試し下さい. "
              )
              return None
      self.parent.withdraw()
      if self.window:
        self.window.lift()
      else:
        self.window = tkinter.Toplevel(self.parent)
        self.window.title("一括採点")
        if func(self, *args, **kargs) is None:
          self.window.protocol("WM_DELETE_WINDOW", self.this_window_close)
          self.window.mainloop()
    return inner

  def check_dir_exist(self):
    self.window.withdraw()
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
    name_project: str = dict_project["name"]
    path_dir    : str = dict_project["path_dir"]
    path_file   : str = dict_project["path_file"]
    if not os.path.exists(path_dir):
      tkinter.messagebox.showwarning(
        "フォルダが存在しません", 
        f"指定されたフォルダが存在しなかったため, フォルダを開くことができませんでした. \n"
        + f"答案スキャンデータが保存されているフォルダのパスが正しいことを確認して下さい. \n\n"
        + f"試験名: {name_project}"
      )
      return False
    if not os.path.exists(path_file):
      tkinter.messagebox.showwarning(
        "ファイルが存在しません", 
        f"指定されたファイルが存在しなかったため, ファイルを開くことができませんでした. \n"
        + f"模範解答スキャンデータが保存されているファイルのパスが正しいことを確認して下さい. \n\n"
        + f"試験名: {name_project}"
      )
      return False
    if not os.path.splitext(path_file)[1] in [".jpeg", ".jpg", ".png"]:
      tkinter.messagebox.showwarning(
        "ファイルの拡張子が対応しません", 
        f"指定されたファイルの拡張子が jpeg, jpg, png 以外であったため, ファイルを開きませんでした. \n"
        + f"模範解答スキャンデータが保存されているファイル名が正しいことを確認して下さい. \n"
        + f"ファイルの形式が正しくない場合は, 外部のアプリケーションを利用してファイルを変換して下さい. \n"
        + f"ファイルの形式が正しい場合は, 拡張子を変更した上でもう一度実行して下さい. \n\n"
        + f"試験名: {name_project}"
      )
      return False
    if not os.path.exists(path_dir + "/.temp_saiten"):
      os.mkdir(path_dir + "/.temp_saiten")
      subprocess.check_call(["attrib", "+H", path_dir + "/.temp_saiten"])
    if not os.path.exists(path_dir + "/.temp_saiten"):
      os.mkdir(path_dir + "/.temp_saiten")
    if not os.path.exists(path_dir + "/.temp_saiten/answer_area.json"):
      dict_answer_area = {"questions": []}
      with open(path_dir + "/.temp_saiten/answer_area.json", "w", encoding="utf-8") as f:
        json.dump(dict_answer_area, f, indent=2)
    if not os.path.exists(path_dir + "/.temp_saiten/model_answer"):
      os.mkdir(path_dir + "/.temp_saiten/model_answer")
    if not os.path.exists(path_dir + "/.temp_saiten/answer"):
      os.mkdir(path_dir + "/.temp_saiten/answer")
    if not os.path.exists(path_dir + "/.temp_saiten/make_xlsx"):
      os.mkdir(path_dir + "/.temp_saiten/make_xlsx")
    if not os.path.exists(path_dir + "/.temp_saiten/model_answer/model_answer.png"):
      if os.path.splitext(path_file)[1] in [".jpeg", ".jpg", ".png"]:
        img = PIL.Image.open(path_file)
        img.save(path_dir + "/.temp_saiten/model_answer/model_answer.png")
    list_path_in_file_dir = [path.replace("\\", "/") for path in glob.glob(path_dir + "/*")]
    if os.path.exists(path_dir + "/.temp_saiten/load_picture.json"):
      with open(path_dir + "/.temp_saiten/load_picture.json", "r", encoding="utf-8") as f:
        dict_load_picture = json.load(f)
    else:
      dict_load_picture = {
        "answer": []
      }
    with open(path_dir + "/.temp_saiten/answer_area.json", "r", encoding="utf-8") as f:
      dict_answer_area = json.load(f)    
    index_file = len(dict_load_picture["answer"])
    for path_in_file_dir in list_path_in_file_dir:
      if path_in_file_dir == path_file:
        continue
      elif path_in_file_dir in dict_load_picture["answer"]:
        continue
      elif os.path.splitext(path_in_file_dir)[1] in [".jpeg", ".jpg", ".png"]:
        img = PIL.Image.open(path_in_file_dir)
        img.save(path_dir + "/.temp_saiten/answer/" + str(index_file) + ".png")
        dict_load_picture["answer"].append(path_in_file_dir)
        for index_questions_score in range(len(dict_answer_area["questions"])):
          dict_answer_area["questions"][index_questions_score]["score"].append({"status": "unscored", "point": None})
      else:
        continue
      index_file += 1
    with open(path_dir + "/.temp_saiten/load_picture.json", "w", encoding="utf-8") as f:
      json.dump(dict_load_picture, f, indent=2)
    with open(path_dir + "/.temp_saiten/answer_area.json", "w", encoding="utf-8") as f:
      json.dump(dict_answer_area, f, indent=2)
    if index_file == 0:
      tkinter.messagebox.showwarning(
        "ファイルが存在しません", 
        f"指定されたフォルダ内に, 拡張子が *.jpeg, *.jpg, *.png であるファイルが存在しません. \n"
        + f"答案スキャンデータが保存されているフォルダ名が正しいことを確認して下さい. \n"
        + f"ファイルの形式が正しくない場合は, 外部のアプリケーションを利用してファイルを変換して下さい. \n"
        + f"ファイルの形式が正しい場合は, 拡張子を変更した上でもう一度実行して下さい. \n\n"
        + f"試験名: {name_project}"
      )
      return False
    tkinter.messagebox.showinfo(
      "答案スキャンデータが読み込みました", 
      f"{index_file} 件の答案スキャンデータが読みこまれています. \n\n"
      + f"読み込まれるスキャンデータが少ない場合は以下の手順で確認して下さい. \n"
      + f"1. メインウインドウの［編集］ボタンをクリックして, 「試験を編集」ウインドウを開きます. \n"
      + f"2. 答案スキャンデータの保存されているフォルダのパスが正しいことを確認して下さい. \n"
      + f"3. 答案スキャンデータとして使用できるファイルは JPEG または PNG です. 拡張子が *.jpeg, *.jpg, *.png 以外のファイルは無視されます. \n"
      + f"4. ［適用］をクリックして, 答案データを再読み込みします. \n\n"
      + f"読み込みには時間がかかる場合があります. 操作をせず10秒程度お待ち下さい. "
    )
    self.window.deiconify()
    return True

  @sub_window_loop
  def add_project(self):
    def choose_dir():
      entry_path_dir.delete(0, "end")
      entry_path_dir.insert(0, tkinter.filedialog.askdirectory())
      self.window.lift()
    
    def choose_file():
      entry_path_file.delete(0, "end")
      entry_path_file.insert(0, tkinter.filedialog.askopenfilename())
      self.window.lift()

    def add_json():
      str_name = entry_name.get()
      str_path_dir = entry_path_dir.get()
      str_path_file = entry_path_file.get()
      if str_name == "":
        tkinter.messagebox.showwarning("試験名が入力されていません", "試験名が入力されていないため, 新しく試験を作成できません. \n試験名を入力して下さい. ")
        self.window.lift()
        return
      if str_path_dir == "":
        tkinter.messagebox.showwarning("フォルダパスが指定されていません", "答案ファイルが保存されているフォルダパスが指定されていないため, 新しく試験を作成できません. \nフォルダパスを指定して下さい. ")
        self.window.lift()
        return
      if str_path_file == "":
        tkinter.messagebox.showwarning("ファイルパスが指定されていません", "模範解答が保存されているファイルパスが指定されていないため, 新しく試験を追加できません. \nファイルパスを指定して下さい. ")
        self.window.lift()
        return
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_config["projects"].append(
        {
          "name": str_name,
          "path_dir": str_path_dir,
          "path_file": str_path_file,
          "export": {
            "symbol": {
              "position": "c",
              "x": 0,
              "y": 0,
              "size": 0,
              "unscored": True,
              "correct": True,
              "partial": True,
              "hold": True,
              "incorrect": True
            },
            "point": {
              "position": "c",
              "x": 0,
              "y": 0,
              "size": 0,
              "unscored": True,
              "correct": True,
              "partial": True,
              "hold": True,
              "incorrect": True
            }
          }
        }
      )
      dict_config["index_projects_in_listbox"] = len(dict_config["projects"]) - 1
      with open("config.json", "w", encoding="utf-8") as f:
        json.dump(dict_config, f, indent=2)
      if not self.check_dir_exist():
        with open("config.json", "r", encoding="utf-8") as f:
          dict_config = json.load(f)
        dict_config["projects"].pop(len(dict_config["projects"]) - 1)
        if len(dict_config["projects"]) == 0:
          dict_config["index_projects_in_listbox"] = None
        else:
          dict_config["index_projects_in_listbox"] = 0
        with open("config.json", "w", encoding="utf-8") as f:
          json.dump(dict_config, f, indent=2)
        self.window.lift()
        return
      self.this_window_close()
      tkinter.messagebox.showinfo(
        "試験を追加しました",
        "採点データ等は, 指定したフォルダ内に作成された隠しフォルダ「.temp_saiten」内に保存されます. \n"
        + "予期せぬ動作を防ぐため, 本アプリ起動中は「.temp_saiten」や指定したフォルダを移動, 削除しないで下さい. "
      )

    self.window.title("試験を追加")
    frame_main = tkinter.Frame(self.window)
    frame_main.pack(expand=True, padx=20, pady=20)

    frame_form = tkinter.Frame(frame_main, width=80, height=10)
    label_vspace = tkinter.Label(frame_main, width=100, height=2)
    frame_btn = tkinter.Frame(frame_main, width=80, height=10)
    frame_form.grid(column=0, row=0)
    label_vspace.grid(column=0, row=1)
    frame_btn.grid(column=0, row=2)
    
    label_name = tkinter.Label(frame_form, width=80, text="試験名")
    label_name.grid(column=0, row=0)
    entry_name = tkinter.Entry(frame_form, width=80)
    entry_name.grid(column=0, row=1)
    label_vspace1 = tkinter.Label(frame_form, width=80, height=1)
    label_vspace1.grid(column=0, row=2)
    label_path_dir = tkinter.Label(frame_form, width=80, text="答案スキャンデータが保存されているフォルダのパス")
    label_path_dir.grid(column=0, row=3)
    frame_path_dir = tkinter.Frame(frame_form, width=80)
    frame_path_dir.grid(column=0, row=4)
    label_vspace2 = tkinter.Label(frame_form, width=80, height=1)
    label_vspace2.grid(column=0, row=5)
    label_path_file = tkinter.Label(frame_form, width=80, text="模範解答スキャンデータが保存されているファイルのパス")
    label_path_file.grid(column=0, row=6)
    frame_path_file = tkinter.Frame(frame_form, width=80)
    frame_path_file.grid(column=0, row=7)

    entry_path_dir = tkinter.Entry(frame_path_dir, width=60, textvariable="")
    entry_path_dir.grid(column=0, row=0)
    label_hspace_dir = tkinter.Label(frame_path_dir, width=3)
    label_hspace_dir.grid(column=1, row=0)
    btn_path_dir = tkinter.Button(frame_path_dir, width=15, text="フォルダを選択", command=choose_dir)
    btn_path_dir.grid(column=2, row=0)

    entry_path_file = tkinter.Entry(frame_path_file, width=60, textvariable="")
    entry_path_file.grid(column=0, row=0)
    label_hspace_file = tkinter.Label(frame_path_file, width=3)
    label_hspace_file.grid(column=1, row=0)
    btn_path_file = tkinter.Button(frame_path_file, width=15, text="模範解答を選択", command=choose_file)
    btn_path_file.grid(column=2, row=0)
    
    tkinter.Button(frame_btn, text="試験を追加", command=add_json, width=40, height=2).grid(column=0, row=0)
    tkinter.Button(frame_btn, text="キャンセル", command=self.this_window_close, width=40, height=2).grid(column=1, row=0)

  def edit_project(self):
    nothing_to_do()
    
  # 解答欄の位置を指定
  @sub_window_loop
  def select_area(self):
    if not self.check_dir_exist():
      tkinter.messagebox.showinfo(
        "設定を確認して下さい", 
        f"試験一覧の［編集］ボタンをクリックして, 試験の設定を確認して下さい. \n\n「解答欄の位置を指定」を終了します. "
      )
      return "break"
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)

    dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
    path_dir = dict_project["path_dir"]
    path_json_answer_area = dict_project["path_dir"] + "/.temp_saiten/answer_area.json"
    path_file_model_answer = dict_project["path_dir"] + "/.temp_saiten/model_answer/model_answer.png"
    path_dir_of_answers = dict_project["path_dir"] + "/.temp_saiten/answer"
    with open(path_json_answer_area, "r", encoding="utf-8") as f:
      dict_answer_area = json.load(f)

    def del_question():
      if self.index_selected_question is not None:
        with open(path_json_answer_area, "r", encoding="utf-8") as f:
          dict_answer_area = json.load(f)
        dict_answer_area["questions"].pop(self.index_selected_question)
        with open(path_json_answer_area, "w", encoding="utf-8") as f:
          json.dump(dict_answer_area, f, indent=2)
        reload_listbox_question()

    def up_question():
      if self.index_selected_question is not None:
        with open(path_json_answer_area, "r", encoding="utf-8") as f:
          dict_answer_area = json.load(f)
        pop_question = dict_answer_area["questions"].pop(self.index_selected_question)
        self.index_selected_question = max(self.index_selected_question - 1, 0)
        dict_answer_area["questions"].insert(self.index_selected_question, pop_question)
        with open(path_json_answer_area, "w", encoding="utf-8") as f:
          json.dump(dict_answer_area, f, indent=2)
        reload_listbox_question()

    def down_question():
      if self.index_selected_question is not None:
        with open(path_json_answer_area, "r", encoding="utf-8") as f:
          dict_answer_area = json.load(f)
        pop_question = dict_answer_area["questions"].pop(self.index_selected_question)
        self.index_selected_question = min(self.index_selected_question + 1, len(dict_answer_area["questions"]) - 1)
        dict_answer_area["questions"].insert(self.index_selected_question, pop_question)
        with open(path_json_answer_area, "w", encoding="utf-8") as f:
          json.dump(dict_answer_area, f, indent=2)
        reload_listbox_question()

    def set_type(str_type):
      if self.index_selected_question is not None:
        with open(path_json_answer_area, "r", encoding="utf-8") as f:
          dict_answer_area = json.load(f)
        dict_answer_area["questions"][self.index_selected_question]["type"] = str_type
        with open(path_json_answer_area, "w", encoding="utf-8") as f:
          json.dump(dict_answer_area, f, indent=2)
        reload_listbox_question()
    def set_question():
      set_type("設問")
    def set_name():
      set_type("氏名")
    def set_id():
      set_type("生徒番号")
    def set_stamp():
      set_type("採点者印")
    def set_subtotal():
      set_type("小計点")
    def set_total():
      set_type("合計点")

    def canvas_draw_rectangle_click(event):
      self.canvas_draw_rectangle[0] = event.x
      self.canvas_draw_rectangle[1] = event.y
      self.canvas_draw_rectangle[2] = min(event.x + 1, canvas.winfo_width())
      self.canvas_draw_rectangle[3] = min(event.y + 1, canvas.winfo_height())
      canvas.coords("rectangle_new",
        self.canvas_draw_rectangle[0],
        self.canvas_draw_rectangle[1],
        self.canvas_draw_rectangle[2],
        self.canvas_draw_rectangle[3], 
      )
    def canvas_draw_rectangle_drag(event):
      self.canvas_draw_rectangle[2] = min(max(event.x, 0), canvas.winfo_width())
      self.canvas_draw_rectangle[3] = min(max(event.y, 0), canvas.winfo_height())
      canvas.coords("rectangle_new",
        self.canvas_draw_rectangle[0],
        self.canvas_draw_rectangle[1],
        self.canvas_draw_rectangle[2],
        self.canvas_draw_rectangle[3], 
      )
    def canvas_draw_rectangle_release(event):
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)
      with open(path_dir + "/.temp_saiten/load_picture.json", "r", encoding="utf-8") as f:
        dict_load_picture = json.load(f)
      dict_answer_area["questions"].append(
        {
          "type": "設問", 
          "daimon": None,
          "shomon": None,
          "shimon": None,
          "haiten": None,
          "area": [
            min(self.canvas_draw_rectangle[0], self.canvas_draw_rectangle[2]),
            min(self.canvas_draw_rectangle[1], self.canvas_draw_rectangle[3]),
            max(self.canvas_draw_rectangle[0], self.canvas_draw_rectangle[2]),
            max(self.canvas_draw_rectangle[1], self.canvas_draw_rectangle[3])
          ],
          "score": [
            {
              "status": "unscored",
              "point": None
            }
            for i in range(len(dict_load_picture["answer"]))
          ]
        }
      )
      with open(path_json_answer_area, "w", encoding="utf-8") as f:
        json.dump(dict_answer_area, f, indent=2)
      self.index_selected_question = len(dict_answer_area["questions"]) - 1
      reload_listbox_question()
      canvas.coords("rectangle_new", 0, 0, 0, 0)

    def selected_listbox_question(*args, **kwargs):
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)      
      for index_question, question in enumerate(dict_answer_area["questions"]):
        if question["type"] == "設問":
          color_reactangle = "green"
        elif question["type"] == "氏名":
          color_reactangle = "blue"
        elif question["type"] == "生徒番号":
          color_reactangle = "cyan"
        elif question["type"] == "小計点":
          color_reactangle = "magenta"
        elif question["type"] == "合計点":
          color_reactangle = "orange"
        elif question["type"] == "採点者印":
          color_reactangle = "yellow"
        self.index_selected_question = listbox_question.curselection()[0]
        if index_question == listbox_question.curselection()[0]:
          color_reactangle = "red"
        canvas.create_rectangle(
          question["area"][0], 
          question["area"][1], 
          question["area"][2], 
          question["area"][3], 
          outline=color_reactangle,
          width=2,
          fill=color_reactangle,
          stipple="gray12",
          tags="field"
        )
        canvas.create_text(
          question["area"][0] - 10, 
          (question["area"][1] + question["area"][3]) // 2, 
          text=str(index_question),
          fill="green",
          tags="number"
        )

    def reload_listbox_question():
      listbox_question.configure(state=tkinter.NORMAL)
      listbox_question.delete(0, tkinter.END)
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)
      canvas.delete("field")
      canvas.delete("number")
      if len(dict_answer_area["questions"]) == 0:
        self.index_selected_question = None
        listbox_question.insert(tkinter.END, "模範解答の画像の上で")
        listbox_question.insert(tkinter.END, "ドラッグして")
        listbox_question.insert(tkinter.END, "解答欄を指定して下さい")
        listbox_question.configure(state=tkinter.DISABLED)
      else:
        self.index_selected_question = min(self.index_selected_question, len(dict_answer_area["questions"]) - 1)
        for index_question, question in enumerate(dict_answer_area["questions"]):
          listbox_question.insert(tkinter.END, f"枠{index_question} - {question['type']}")
        listbox_question.select_set(self.index_selected_question)
        selected_listbox_question()

    self.window.title("解答欄を指定")
    self.canvas_draw_rectangle = [0, 0, 0, 0]

    frame_main = tkinter.Frame(self.window)
    frame_main.grid(column=0, row=0)

    frame_question = tkinter.Frame(frame_main)
    frame_question.grid(column=0, row=0)
    frame_picture = tkinter.Frame(frame_main)
    frame_picture.grid(column=1, row=0)

    frame_listbox_question = tkinter.Frame(frame_question)
    frame_listbox_question.grid(column=0, row=0)
    frame_btn_list_question = tkinter.Frame(frame_question)
    frame_btn_list_question.grid(column=0, row=1)

    listbox_question = tkinter.Listbox(frame_listbox_question, width=20, height=30)
    listbox_question.pack(side="left")
    listbox_question.configure(
      activestyle=tkinter.DOTBOX,
      selectmode=tkinter.SINGLE,
      selectbackground="grey"
    )
    for index_question in range(len(dict_answer_area["questions"])):
      listbox_question.insert(tkinter.END, f"設問{index_question}")   
    listbox_question.bind("<MouseWheel>", lambda eve:listbox_question.yview_scroll(int(-eve.delta/120), 'units'))
    yscrollbar_table_question = tkinter.Scrollbar(frame_listbox_question, orient=tkinter.VERTICAL, command=listbox_question.yview)
    yscrollbar_table_question.pack(side="right", fill="y")
    listbox_question.config(
      yscrollcommand=yscrollbar_table_question.set
    )
    
    btn_list_question_del = tkinter.Button(frame_btn_list_question, width=6, text="削除", command=del_question)
    btn_list_question_del.grid(column=0, row=0)
    btn_list_question_up = tkinter.Button(frame_btn_list_question, width=6, text="上へ", command=up_question)
    btn_list_question_up.grid(column=1, row=0)
    btn_list_question_down = tkinter.Button(frame_btn_list_question, width=6, text="下へ", command=down_question)
    btn_list_question_down.grid(column=2, row=0)
    btn_list_question_que = tkinter.Button(frame_btn_list_question, width=6, text="設問", command=set_question)
    btn_list_question_que.grid(column=0, row=1)
    btn_list_question_name = tkinter.Button(frame_btn_list_question, width=6, text="氏名", command=set_name)
    btn_list_question_name.grid(column=1, row=1)
    btn_list_question_id = tkinter.Button(frame_btn_list_question, width=6, text="生徒番号", command=set_id)
    btn_list_question_id.grid(column=2, row=1)
    btn_list_question_id = tkinter.Button(frame_btn_list_question, width=6, text="採点者印", command=set_stamp)
    btn_list_question_id.grid(column=0, row=2)
    btn_list_question_subtotal = tkinter.Button(frame_btn_list_question, width=6, text="小計点", command=set_subtotal)
    btn_list_question_subtotal.grid(column=1, row=2)
    btn_list_question_total = tkinter.Button(frame_btn_list_question, width=6, text="合計点", command=set_total)
    btn_list_question_total.grid(column=2, row=2)
    btn_scale_up = tkinter.Button(frame_btn_list_question, width=6, text="拡大")
    btn_scale_up.grid(column=0, row=3)
    btn_scale_reset = tkinter.Button(frame_btn_list_question, width=6, text="100%")
    btn_scale_reset.grid(column=1, row=3)
    btn_scale_down = tkinter.Button(frame_btn_list_question, width=6, text="縮小")
    btn_scale_down.grid(column=2, row=3)
    btn_scale_mode = tkinter.Button(frame_btn_list_question, width=21, text="[ドラッグ] / 自動")
    btn_scale_mode.grid(column=0, row=4, columnspan=3)
    btn_scale_help = tkinter.Button(frame_btn_list_question, width=21, text="ヘルプ")
    btn_scale_help.grid(column=0, row=5, columnspan=3)
    btn_scale_back = tkinter.Button(frame_btn_list_question, width=21, text="戻る", command=self.this_window_close)
    btn_scale_back.grid(column=0, row=6, columnspan=3)

    frame_canvas = tkinter.Frame(frame_picture)
    frame_canvas.pack()

    canvas = tkinter.Canvas(frame_canvas, bg="black", width=567, height=800)
    canvas.bind("<Control-MouseWheel>", lambda eve:canvas.xview_scroll(int(-eve.delta/120), 'units'))
    canvas.bind("<MouseWheel>", lambda eve:canvas.yview_scroll(int(-eve.delta/120), 'units'))
    self.tk_image_model_answer = PIL.ImageTk.PhotoImage(file=path_file_model_answer)
    canvas.create_image(0, 0, image=self.tk_image_model_answer, anchor="nw")
    yscrollbar_canvas = tkinter.Scrollbar(frame_canvas, orient=tkinter.VERTICAL, command=canvas.yview)
    xscrollbar_canvas = tkinter.Scrollbar(frame_canvas, orient=tkinter.HORIZONTAL, command=canvas.xview)
    yscrollbar_canvas.pack(side="right", fill="y")
    xscrollbar_canvas.pack(side="bottom", fill="x")
    canvas.pack()
    canvas.config(
      xscrollcommand=xscrollbar_canvas.set,
      yscrollcommand=yscrollbar_canvas.set,
      scrollregion=(0, 0, self.tk_image_model_answer.width(), self.tk_image_model_answer.height())
    )
    
    listbox_question.bind("<<ListboxSelect>>", selected_listbox_question)
    canvas.coords("rectangle_new", 0, 0, 0, 0)
    canvas.create_rectangle(0, 0, 0, 0, fill="red", tags="rectangle_new")
    canvas.bind("<Button-1>", canvas_draw_rectangle_click)
    canvas.bind("<B1-Motion>", canvas_draw_rectangle_drag)
    canvas.bind("<ButtonRelease-1>", canvas_draw_rectangle_release)
    
    if len(dict_answer_area["questions"]) == 0:
      self.index_selected_question = None
    else:
      self.index_selected_question = len(dict_answer_area["questions"]) - 1

    reload_listbox_question()

  @sub_window_loop
  def score_answer(self):
    def help_score_answer(**kwargs):
      tkinter.messagebox.showinfo(
        "使い方",
        "［解答欄の位置を指定］で指定した解答欄ごとに各答案用紙が切り取られ, 設問ごとに採点することができます. \n"
        + "採点する設問は左の設問一覧から選びます. \n\n"
        + "水色 (cyan) で塗られた答案用紙が現在選択されています. \n"
        + "この状態で, ［E］を押すとこの答案のこの設問は「正答」となり, 採点データが保存されます. \n"
        + "採点すると自動的に次の問題が選択され, 採点を続けることができます. \n"
        + "ページに表示されている全部の答案の採点が終わったら［R］を押して答案を再読み込みします. \n"
        + "「表示する答案を選択：」でチェックボックスにチェックが入っている条件で再読み込みされ, 全ての答案の採点が終わるまで繰り返します. \n\n"
        + "選択されている答案は WASD キーで変更できます. \n"
        + "誤った採点を上書きするとき等にお使い下さい. \n\n"
        + "数字キー (0, 1, …) を押すと, 部分点採点モードになり, 部分点として記録できます. \n"
        + "BackSpace キーを押すと, 部分点を削除できます. \n"
        + "［F］または［J］で「部分点」または「保留」として登録して下さい. \n"
        + "採点基準が曖昧である等、後から一括で再採点したい場合等に「保留」をお使い下さい. \n\n"
        + "未採点, 正答, 誤答のいずれかとして採点すると, 部分点情報は削除されます. 予めご了承下さい. "
      )

    if not self.check_dir_exist():
      tkinter.messagebox.showinfo(
        "設定を確認して下さい", 
        f"試験一覧の［編集］ボタンをクリックして, 試験の設定を確認して下さい. \n\n「解答欄の位置を指定」を終了します. "
      )
      return "break"
    self.parent.winfo_screenwidth()
    self.window.geometry("1600x1000+0+0")
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
    path_dir = dict_project["path_dir"]
    path_json_answer_area = dict_project["path_dir"] + "/.temp_saiten/answer_area.json"
    path_file_model_answer = dict_project["path_dir"] + "/.temp_saiten/model_answer/model_answer.png"
    path_dir_of_answers = dict_project["path_dir"] + "/.temp_saiten/answer"
    with open(path_json_answer_area, "r", encoding="utf-8") as f:
      dict_answer_area = json.load(f)
    list_path_file_answer = glob.glob(path_dir_of_answers + "/*")
    
    width_window = self.window.winfo_width()
    height_window = self.window.winfo_height()

    frame_list_question = tkinter.Frame(self.window, padx=10, pady=10, borderwidth=5)
    frame_list_question.grid(column=0, row=0)
    frame_score_question = tkinter.Frame(self.window, padx=10, pady=10)
    frame_score_question.grid(column=1, row=0, sticky=tkinter.NW)

    label_list_question = tkinter.Label(frame_list_question, text="設問一覧", height=2)
    label_list_question.grid(column=0, row=0)
    listbox_question = tkinter.Listbox(frame_list_question)
    listbox_question.grid(column=0, row=1)
    btn_help = tkinter.Button(frame_list_question, width=20, text="ヘルプ", command=help_score_answer)
    btn_help.grid(column=0, row=2)
    btn_quit = tkinter.Button(frame_list_question, width=20, text="戻る", command=self.this_window_close)
    btn_quit.grid(column=0, row=3)

    frame_btn_operate = tkinter.Frame(frame_score_question, background="#bfbfbf")
    frame_btn_operate.grid(column=0, row=0, sticky="we")
    frame_list_frame_canvas_answer = tkinter.Frame(frame_score_question)
    frame_list_frame_canvas_answer.grid(column=0, row=1)

    frame_bar_top = tkinter.Frame(frame_btn_operate, height=5, background="#bfbfbf")
    frame_bar_top.grid(column=0, row=0, sticky="we")
    frame_btn_scoring = tkinter.Frame(frame_btn_operate, height=5)
    frame_btn_scoring.grid(column=0, row=1, padx=5, sticky="w")
    frame_bar_bottom = tkinter.Frame(frame_btn_operate, height=5, background="#bfbfbf")
    frame_bar_bottom.grid(column=0, row=2, sticky="we")

    frame_label_btn_scoring = tkinter.Label(frame_btn_scoring, width=12, text="採点する：")
    frame_border_btn_scoring_unscored = tkinter.Frame(frame_btn_scoring, background="gray")
    frame_border_btn_scoring_correct = tkinter.Frame(frame_btn_scoring, background="green")
    frame_border_btn_scoring_partial = tkinter.Frame(frame_btn_scoring, background="orange")
    frame_border_btn_scoring_hold = tkinter.Frame(frame_btn_scoring, background="blue")
    frame_border_btn_scoring_incorrect = tkinter.Frame(frame_btn_scoring, background="red")
    frame_label_btn_scoring.grid(column=0, row=0)
    frame_border_btn_scoring_unscored.grid(column=1, row=0)
    frame_border_btn_scoring_correct.grid(column=2, row=0)
    frame_border_btn_scoring_partial.grid(column=3, row=0)
    frame_border_btn_scoring_hold.grid(column=4, row=0)
    frame_border_btn_scoring_incorrect.grid(column=5, row=0)
    btn_scoring_unscored = tkinter.Button(frame_border_btn_scoring_unscored, width=15, text="未採点 (Q) ")
    btn_scoring_correct = tkinter.Button(frame_border_btn_scoring_correct, width=15, text="正答 (E) ")
    btn_scoring_partial = tkinter.Button(frame_border_btn_scoring_partial, width=15, text="部分点 (F) ")
    btn_scoring_hold = tkinter.Button(frame_border_btn_scoring_hold, width=15, text="保留 (J) ")
    btn_scoring_incorrect = tkinter.Button(frame_border_btn_scoring_incorrect, width=15, text="誤答 (O) ")
    btn_scoring_unscored.pack(padx=4, pady=4)
    btn_scoring_correct.pack(padx=4, pady=4)
    btn_scoring_partial.pack(padx=4, pady=4)
    btn_scoring_hold.pack(padx=4, pady=4)
    btn_scoring_incorrect.pack(padx=4, pady=4)

    frame_bar = tkinter.Frame(frame_btn_scoring, height=5, background="#bfbfbf")
    frame_bar.grid(column=0, row=1, columnspan=6, sticky="we")
    
    self.booleanVar_checkbutton_show = {
      "unscored": tkinter.BooleanVar(value=True),
      "correct": tkinter.BooleanVar(value=False),
      "partial": tkinter.BooleanVar(value=False),
      "hold": tkinter.BooleanVar(value=False),
      "incorrect": tkinter.BooleanVar(value=False),
    }
    frame_label_checkbotton_show = tkinter.Label(frame_btn_scoring, width=12, text="表示する\n答案を選択：")
    frame_border_checkbutton_show_unscored = tkinter.Frame(frame_btn_scoring, background="gray")
    frame_border_checkbutton_show_correct = tkinter.Frame(frame_btn_scoring, background="green")
    frame_border_checkbutton_show_partial = tkinter.Frame(frame_btn_scoring, background="orange")
    frame_border_checkbutton_show_hold = tkinter.Frame(frame_btn_scoring, background="blue")
    frame_border_checkbutton_show_incorrect = tkinter.Frame(frame_btn_scoring, background="red")
    frame_label_checkbotton_show.grid(column=0, row=2, sticky="we")
    frame_border_checkbutton_show_unscored.grid(column=1, row=2, sticky="we")
    frame_border_checkbutton_show_correct.grid(column=2, row=2, sticky="we")
    frame_border_checkbutton_show_partial.grid(column=3, row=2, sticky="we")
    frame_border_checkbutton_show_hold.grid(column=4, row=2, sticky="we")
    frame_border_checkbutton_show_incorrect.grid(column=5, row=2, sticky="we")
    checkbutton_show_unscored = tkinter.Checkbutton(
      frame_border_checkbutton_show_unscored, 
      width=12, 
      text="未採点 (Ctrl + Q) ",
      variable=self.booleanVar_checkbutton_show["unscored"]
    )
    checkbutton_show_correct = tkinter.Checkbutton(
      frame_border_checkbutton_show_correct, 
      width=12, 
      text="正答 (Ctrl + E) ",
      variable=self.booleanVar_checkbutton_show["correct"]
    )
    checkbutton_show_partial = tkinter.Checkbutton(
      frame_border_checkbutton_show_partial, 
      width=12, 
      text="部分点 (Ctrl + F) ",
      variable=self.booleanVar_checkbutton_show["partial"]
    )
    checkbutton_show_hold = tkinter.Checkbutton(
      frame_border_checkbutton_show_hold, 
      width=12, 
      text="保留 (Ctrl + J) ",
      variable=self.booleanVar_checkbutton_show["hold"]
    )
    checkbutton_show_incorrect = tkinter.Checkbutton(
      frame_border_checkbutton_show_incorrect, 
      width=12, 
      text="誤答 (Ctrl + O) ",
      variable=self.booleanVar_checkbutton_show["incorrect"]
    )
    checkbutton_show_unscored.pack(padx=4, pady=4)
    checkbutton_show_correct.pack(padx=4, pady=4)
    checkbutton_show_partial.pack(padx=4, pady=4)
    checkbutton_show_hold.pack(padx=4, pady=4)
    checkbutton_show_incorrect.pack(padx=4, pady=4)
    checkbutton_show_unscored.pack(padx=4, pady=4)


    self.scoring_model_images = PIL.ImageTk.PhotoImage(file=path_file_model_answer)
    # self.list_path_file_answer = []
    self.list_scoring_images = []
    for path_file_answer in list_path_file_answer:
      if os.path.splitext(path_file_answer)[1] == ".png":
        # self.list_path_file_answer.append(path_file_answer)
        self.list_scoring_images.append(PIL.ImageTk.PhotoImage(file=path_file_answer))
    
    self.index_selected_scoring_question = 0
    self.relation_index_scoring_question_to_index_question = []
    for index_question, question in enumerate(dict_answer_area["questions"]):
      if question["type"] == "設問":
        name_question = "設問"
        if question["daimon"] is not None:
          name_question += " - " + str(question["daimon"])
        if question["shomon"] is not None:
          name_question += " - " + str(question["shomon"])
        if question["shimon"] is not None:
          name_question += " - " + str(question["shimon"])
        listbox_question.insert(tkinter.END, name_question)
        self.relation_index_scoring_question_to_index_question.append(index_question)
    listbox_question.select_set(self.index_selected_scoring_question)
    self.index_selected_scoring_question = self.relation_index_scoring_question_to_index_question[0]

    def repack_chosen_frame_canvas_answer(self):
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)
      label_show_page.configure(text=f"{self.index_pages_relation_table_position_to_index_answersheet + 1} 頁 / {len(self.pages_relation_table_position_to_index_answersheet)} 頁")
      for index_relation_table_position_to_index_answersheet, ((int_column_position_of_answer, int_row_position_of_answer), index_scoring_answersheet) in enumerate(self.pages_relation_table_position_to_index_answersheet[
        self.index_pages_relation_table_position_to_index_answersheet]):
        self.list_entry_score[index_scoring_answersheet].configure(state="normal")
        if dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["status"] == "unscored":
          background_frame = "gray"
          self.list_entry_score[index_scoring_answersheet].delete(0, tkinter.END)
          self.list_entry_score[index_scoring_answersheet].insert(0, "未採")
        elif dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["status"] == "correct":
          background_frame = "green"
          self.list_entry_score[index_scoring_answersheet].delete(0, tkinter.END)
          if dict_answer_area["questions"][self.index_selected_scoring_question]["haiten"] is None:
            self.list_entry_score[index_scoring_answersheet].insert(0, "配")
          else:
            self.list_entry_score[index_scoring_answersheet].insert(0, str(dict_answer_area["questions"][self.index_selected_scoring_question]["haiten"]))
        elif dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["status"] == "partial":
          background_frame = "yellow"
          self.list_entry_score[index_scoring_answersheet].delete(0, tkinter.END)
          if dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["point"] is not None:
            self.list_entry_score[index_scoring_answersheet].insert(0, str(dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["point"]))
        elif dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["status"] == "hold":
          background_frame = "blue"
          self.list_entry_score[index_scoring_answersheet].delete(0, tkinter.END)
          if dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["point"] is not None:
            self.list_entry_score[index_scoring_answersheet].insert(0, str(dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["point"]))
        elif dict_answer_area["questions"][self.index_selected_scoring_question]["score"][index_scoring_answersheet]["status"] == "incorrect":
          background_frame = "red"
          self.list_entry_score[index_scoring_answersheet].configure(state="normal")
          self.list_entry_score[index_scoring_answersheet].delete(0, tkinter.END)
          self.list_entry_score[index_scoring_answersheet].insert(0, "0")
        self.list_entry_score[index_scoring_answersheet].configure(state="readonly")
        self.list_frame_border_frame_canvas_question[index_scoring_answersheet].configure(background=background_frame)
        self.list_frame_border_frame_canvas_question[index_scoring_answersheet].grid(column=int_column_position_of_answer , row=int_row_position_of_answer, padx=2, pady=2)
        self.list_frame_canvas_question[index_scoring_answersheet].configure(background="white")
        self.list_label_entry_score[index_scoring_answersheet].configure(background="white")
        if index_relation_table_position_to_index_answersheet == self.index_selected_relation_table_position_to_index_answersheet:
          self.list_frame_canvas_question[index_scoring_answersheet].configure(background="cyan")
          self.list_label_entry_score[index_scoring_answersheet].configure(background="cyan")
        self.list_frame_canvas_question[index_scoring_answersheet].grid(padx=3, pady=3)
        self.list_canvas_question[index_scoring_answersheet].grid(column=0, row=0, columnspan=2, padx=1, pady=1)
        self.list_entry_score[index_scoring_answersheet].grid(column=0, row=1, sticky="e")
        self.list_label_entry_score[index_scoring_answersheet].grid(column=1, row=1, sticky="w")

    def choose_to_show_frame_canvas_answer(self):
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
      path_dir = dict_project["path_dir"]
      path_json_answer_area = path_dir + "/.temp_saiten/answer_area.json"
      path_dir_of_answers = path_dir + "/.temp_saiten/answer"
      path_file_model_answer = path_dir + "/.temp_saiten/model_answer/model_answer.png"
      list_path_file_answer = glob.glob(path_dir_of_answers + "/*")
      path_dir_of_answers = path_dir + "/.temp_saiten/answer"
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)

      if self.index_selected_scoring_question is None:
        self.index_pages_relation_table_position_to_index_answersheet = None
      else:
        self.window.update_idletasks()
        width_window = self.window.winfo_width()
        height_window = self.window.winfo_height()

        listbox_question.configure(height=height_window // 21 - 5)
        frame_list_question.update_idletasks()
        frame_btn_operate.update_idletasks()
        width_frame_list_question = frame_list_question.winfo_width()
        height_frame_btn_operate = frame_btn_operate.winfo_height()

        width_canvas = dict_answer_area["questions"][self.index_selected_scoring_question]["area"][2] - dict_answer_area["questions"][self.index_selected_scoring_question]["area"][0]
        height_canvas = dict_answer_area["questions"][self.index_selected_scoring_question]["area"][3] - dict_answer_area["questions"][self.index_selected_scoring_question]["area"][1]      

        self.len_column_position_of_answer = (width_window - width_frame_list_question) // (width_canvas + 20)
        self.len_row_position_of_answer = (height_window - 150) // (height_canvas + 40)

        self.frame_border_frame_canvas_model_answer.grid(column=0, row=0)
        self.frame_canvas_model_answer.grid(padx=4, pady=4)
        self.canvas_model_answer.grid(column=0, row=0)
        self.label_model_answer.grid(column=0, row=1)
        
        int_column_position_of_answer = 1
        int_row_position_of_answer = 0

        self.pages_relation_table_position_to_index_answersheet = [[]]
        for index_scoring_answersheet, scoring_answersheet in enumerate(dict_answer_area["questions"][self.index_selected_scoring_question]["score"]):
          if self.booleanVar_checkbutton_show[scoring_answersheet["status"]].get():
            self.pages_relation_table_position_to_index_answersheet[-1].append(((int_column_position_of_answer, int_row_position_of_answer), index_scoring_answersheet))
            int_column_position_of_answer += 1
            if int_column_position_of_answer == self.len_column_position_of_answer:
              int_column_position_of_answer = 0
              int_row_position_of_answer += 1
            if int_row_position_of_answer == self.len_row_position_of_answer and index_scoring_answersheet != len(dict_answer_area["questions"][self.index_selected_scoring_question]["score"]) - 1:
              self.pages_relation_table_position_to_index_answersheet.append([])
              int_column_position_of_answer = 1
              int_row_position_of_answer = 0
        self.index_selected_relation_table_position_to_index_answersheet = 0
        repack_chosen_frame_canvas_answer(self)

    def reload_frame_canvas_answer(self, *args, **kwargs):
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
      path_dir = dict_project["path_dir"]
      path_json_answer_area = path_dir + "/.temp_saiten/answer_area.json"
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)

      frame_list_frame_canvas_answer.grid_forget()
      frame_list_frame_canvas_answer.grid(column=0, row=1, sticky="nw")
      self.frame_border_frame_canvas_model_answer.destroy()
      for canvas_question in self.list_frame_border_frame_canvas_question:
        canvas_question.destroy()
      self.list_frame_border_frame_canvas_question = []
      self.list_frame_canvas_question = []
      self.list_canvas_question = []
      self.list_label_entry_score = []
      self.list_entry_score = []
      width_canvas = dict_answer_area["questions"][self.index_selected_scoring_question]["area"][2] - dict_answer_area["questions"][self.index_selected_scoring_question]["area"][0]
      height_canvas = dict_answer_area["questions"][self.index_selected_scoring_question]["area"][3] - dict_answer_area["questions"][self.index_selected_scoring_question]["area"][1]
      
      self.frame_border_frame_canvas_model_answer = tkinter.Frame(frame_list_frame_canvas_answer, background="black")
      self.frame_canvas_model_answer = tkinter.Frame(self.frame_border_frame_canvas_model_answer)
      self.canvas_model_answer = tkinter.Canvas(self.frame_canvas_model_answer, width=width_canvas, height=height_canvas)
      self.canvas_model_answer.create_image(
        -1 * dict_answer_area["questions"][self.index_selected_scoring_question]["area"][0],
        -1 * dict_answer_area["questions"][self.index_selected_scoring_question]["area"][1],
        image=self.scoring_model_images,
        anchor="nw",
        tags="answer"
      )
      self.label_model_answer = tkinter.Label(self.frame_canvas_model_answer)
      if dict_answer_area['questions'][self.index_selected_scoring_question]['haiten'] is None:
        self.label_model_answer.configure(text=f"模範解答: 未配点")
      else:
        self.label_model_answer.configure(text=f"模範解答: {dict_answer_area['questions'][self.index_selected_scoring_question]['haiten']}点")
      
      if len(dict_answer_area["questions"][self.index_selected_scoring_question]) == 0:
        self.index_selected_column_position_of_answer = None
        self.index_selected_row_position_of_answer = None
        self.index_selected_column_position_of_answer = None
      else:
        self.index_selected_column_position_of_answer = 1
        self.index_selected_row_position_of_answer = 0
        self.index_selected_scoring_answersheet = 0
      self.index_pages_relation_table_position_to_index_answersheet = 0

      for index_scoring_answersheet, scoring_answersheet in enumerate(dict_answer_area["questions"][self.index_selected_scoring_question]["score"]):
        self.list_frame_border_frame_canvas_question.append(tkinter.Frame(frame_list_frame_canvas_answer)) #, background=background_frame))
        self.list_frame_canvas_question.append(tkinter.Frame(self.list_frame_border_frame_canvas_question[-1]))
        self.list_canvas_question.append(tkinter.Canvas(self.list_frame_canvas_question[-1], width=width_canvas, height=height_canvas))
        self.list_canvas_question[index_scoring_answersheet].create_image(
          -1 * dict_answer_area["questions"][self.index_selected_scoring_question]["area"][0], 
          -1 * dict_answer_area["questions"][self.index_selected_scoring_question]["area"][1], 
          image=self.list_scoring_images[index_scoring_answersheet], 
          anchor="nw",
          tags="answer"
        )
        self.list_entry_score.append(tkinter.Entry(self.list_frame_canvas_question[-1], width=5, justify="right"))
        self.list_label_entry_score.append(tkinter.Label(self.list_frame_canvas_question[-1], width=3, text="点", justify="left"))
      
      choose_to_show_frame_canvas_answer(self)

    def selected_scoring_question(*args, **kwargs):
      self.index_selected_scoring_question = self.relation_index_scoring_question_to_index_question[listbox_question.curselection()[0]]
      reload_frame_canvas_answer(self)

    def move_selected_question_answersheet(direction: str, *args, **kwargs):
      if len(self.pages_relation_table_position_to_index_answersheet[0]) > 0:
        if direction in ["up", "down", "next", "back"]:
          if direction == "up":
            if self.index_selected_relation_table_position_to_index_answersheet == 0:
              self.index_selected_relation_table_position_to_index_answersheet = len(self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet]) - 1
            else:
              self.index_selected_relation_table_position_to_index_answersheet -= self.len_column_position_of_answer
              if self.index_selected_relation_table_position_to_index_answersheet < 0:
                self.index_selected_relation_table_position_to_index_answersheet = 0
          elif direction == "down":
            if self.index_selected_relation_table_position_to_index_answersheet == len(self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet]) - 1:
              self.index_selected_relation_table_position_to_index_answersheet = 0
            else:
              self.index_selected_relation_table_position_to_index_answersheet += self.len_column_position_of_answer
              if self.index_selected_relation_table_position_to_index_answersheet > len(self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet]) - 1:
                self.index_selected_relation_table_position_to_index_answersheet = len(self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet]) - 1
          elif direction == "next":
            self.index_selected_relation_table_position_to_index_answersheet += 1
            if self.index_selected_relation_table_position_to_index_answersheet == len(self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet]):
              self.index_selected_relation_table_position_to_index_answersheet = 0
          elif direction == "back":
            self.index_selected_relation_table_position_to_index_answersheet -= 1
            if self.index_selected_relation_table_position_to_index_answersheet == -1:
              self.index_selected_relation_table_position_to_index_answersheet = len(self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet]) - 1
          self.index_selected_scoring_answersheet = self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet][self.index_selected_relation_table_position_to_index_answersheet][1]
          repack_chosen_frame_canvas_answer(self)
        else:
          for index_relation_table_position_to_index_answersheet, ((int_column_position_of_answer, int_row_position_of_answer), index_scoring_answersheet) in enumerate(self.pages_relation_table_position_to_index_answersheet[
            self.index_pages_relation_table_position_to_index_answersheet]):
            self.list_frame_border_frame_canvas_question[index_scoring_answersheet].grid_forget()  
            self.list_frame_canvas_question[index_scoring_answersheet].grid_forget()
            self.list_canvas_question[index_scoring_answersheet].grid_forget()
            self.list_entry_score[index_scoring_answersheet].grid_forget()
            self.list_label_entry_score[index_scoring_answersheet].grid_forget()
          if direction == "page_back":
            if self.index_pages_relation_table_position_to_index_answersheet > 0:
              self.index_pages_relation_table_position_to_index_answersheet -= 1
              self.index_selected_relation_table_position_to_index_answersheet = 0
          elif direction == "page_next":
            if self.index_pages_relation_table_position_to_index_answersheet < len(self.pages_relation_table_position_to_index_answersheet) - 1:
              self.index_pages_relation_table_position_to_index_answersheet += 1
              self.index_selected_relation_table_position_to_index_answersheet = 0
          self.index_selected_scoring_answersheet = self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet][self.index_selected_relation_table_position_to_index_answersheet][1]
          self.index_selected_relation_table_position_to_index_answersheet = 0
          repack_chosen_frame_canvas_answer(self)

    def score_selected_question_answersheet(value: str, event):
      self.index_selected_scoring_answersheet = self.pages_relation_table_position_to_index_answersheet[self.index_pages_relation_table_position_to_index_answersheet][self.index_selected_relation_table_position_to_index_answersheet][1]
      if self.index_selected_scoring_answersheet is not None:
        with open(path_json_answer_area, "r", encoding="utf-8") as f:
          dict_answer_area = json.load(f)
        if value in ["unscored", "correct", "partial", "hold", "incorrect"]:
          dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["status"] = value
        else:
          dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["status"] = "partial"
        if value in ["unscored", "correct", "incorrect"]:
          dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["point"] = None
        elif value in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
          if dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["point"] is None:
            dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["point"] = int(value)
          else:
            dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["point"] *= 10
            dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["point"] += int(value)
        elif value in ["backspace"]:
          dict_answer_area["questions"][self.index_selected_scoring_question]["score"][self.index_selected_scoring_answersheet]["point"] = None
        with open(path_json_answer_area, "w", encoding="utf-8") as f:
          json.dump(dict_answer_area, f, indent=2)
        if value in ["unscored", "correct", "partial", "hold", "incorrect"]:
          move_selected_question_answersheet("next")
        else:
          repack_chosen_frame_canvas_answer(self)

    frame_border_btn_reload_answer = tkinter.Frame(frame_btn_scoring, background="cyan")
    frame_border_btn_reload_answer.grid(column=6, row=0)
    btn_reload_answer = tkinter.Button(frame_border_btn_reload_answer, width=15, height=1, text="再読み込み (R)", command=functools.partial(reload_frame_canvas_answer, self))
    btn_reload_answer.grid(column=0, row=0, padx=4, pady=4, sticky="wens")
    frame_bar_between_btn_reload_and_show_page = tkinter.Frame(frame_btn_scoring, background="gray")
    frame_bar_between_btn_reload_and_show_page.grid(column=6, row=1, sticky="wens")
    frame_border_label_show_page = tkinter.Frame(frame_btn_scoring, background="gray")
    frame_border_label_show_page.grid(column=6, row=2, padx=4, pady=4)
    label_show_page = tkinter.Label(frame_border_label_show_page, width=15)
    label_show_page.grid(column=0, row=0)

    frame_border_btn_move_answer_page_back = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_up = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_page_next = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_bar = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_back = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_down = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_next = tkinter.Frame(frame_btn_scoring, background="black")
    frame_border_btn_move_answer_page_back.grid(column=7, row=0)
    frame_border_btn_move_answer_up.grid(column=8, row=0)
    frame_border_btn_move_answer_page_next.grid(column=9, row=0)
    frame_border_btn_move_answer_bar.grid(column=7, row=1, columnspan=3, sticky="wens")
    frame_border_btn_move_answer_back.grid(column=7, row=2)
    frame_border_btn_move_answer_down.grid(column=8, row=2)
    frame_border_btn_move_answer_next.grid(column=9, row=2)
    btn_move_answer_page_back = tkinter.Button(frame_border_btn_move_answer_page_back, width=12, text="前頁 (Shift + A)", command=functools.partial(move_selected_question_answersheet, "page_back"))
    btn_move_answer_up = tkinter.Button(frame_border_btn_move_answer_up, width=12, text="上へ (W)", command=functools.partial(move_selected_question_answersheet, "up"))
    btn_move_answer_page_next = tkinter.Button(frame_border_btn_move_answer_page_next, width=12, text="後頁 (Shift + D)", command=functools.partial(move_selected_question_answersheet, "page_next"))
    btn_move_answer_back = tkinter.Button(frame_border_btn_move_answer_back, width=12, text="左へ (A)", command=functools.partial(move_selected_question_answersheet, "back"))
    btn_move_answer_down = tkinter.Button(frame_border_btn_move_answer_down, width=12, text="下へ (S)", command=functools.partial(move_selected_question_answersheet, "down"))
    btn_move_answer_next = tkinter.Button(frame_border_btn_move_answer_next, width=12, text="右へ (D)", command=functools.partial(move_selected_question_answersheet, "next"))
    btn_move_answer_page_back.grid(column=0, row=0, padx=4, pady=4)
    btn_move_answer_up.grid(column=0, row=0, padx=4, pady=4)
    btn_move_answer_page_next.grid(column=0, row=0, padx=4, pady=4)
    btn_move_answer_back.grid(column=0, row=0, padx=4, pady=4)
    btn_move_answer_down.grid(column=0, row=0, padx=4, pady=4)
    btn_move_answer_next.grid(column=0, row=0, padx=4, pady=4)
    listbox_question.bind("<<ListboxSelect>>", selected_scoring_question)
    
    self.frame_border_frame_canvas_model_answer = tkinter.Frame(frame_list_frame_canvas_answer, background="black")
    self.list_frame_border_frame_canvas_question = []
    self.list_canvas_question = []

    
    self.window.bind("r", functools.partial(reload_frame_canvas_answer, self))
    reload_frame_canvas_answer(self)

    def toggle_booleanVar_checkbutton_show(status, event):
      self.booleanVar_checkbutton_show[status].set(not self.booleanVar_checkbutton_show[status].get())
      choose_to_show_frame_canvas_answer(self)

    self.window.bind("w", functools.partial(move_selected_question_answersheet, "up")) # 上へ
    self.window.bind("s", functools.partial(move_selected_question_answersheet, "down")) # 下へ
    self.window.bind("a", functools.partial(move_selected_question_answersheet, "back")) # 右へ
    self.window.bind("d", functools.partial(move_selected_question_answersheet, "next")) # 左へ
    self.window.bind("A", functools.partial(move_selected_question_answersheet, "page_back")) # 右へ
    self.window.bind("D", functools.partial(move_selected_question_answersheet, "page_next")) # 左へ
    self.window.bind("q", functools.partial(score_selected_question_answersheet, "unscored")) # 未採点
    self.window.bind("e", functools.partial(score_selected_question_answersheet, "correct")) # 正答
    self.window.bind("f", functools.partial(score_selected_question_answersheet, "partial")) # 部分点
    self.window.bind("j", functools.partial(score_selected_question_answersheet, "hold")) # 保留
    self.window.bind("o", functools.partial(score_selected_question_answersheet, "incorrect")) # 誤答
    self.window.bind("0", functools.partial(score_selected_question_answersheet, "0"))
    self.window.bind("1", functools.partial(score_selected_question_answersheet, "1"))
    self.window.bind("2", functools.partial(score_selected_question_answersheet, "2"))
    self.window.bind("3", functools.partial(score_selected_question_answersheet, "3"))
    self.window.bind("4", functools.partial(score_selected_question_answersheet, "4"))
    self.window.bind("5", functools.partial(score_selected_question_answersheet, "5"))
    self.window.bind("6", functools.partial(score_selected_question_answersheet, "6"))
    self.window.bind("7", functools.partial(score_selected_question_answersheet, "7"))
    self.window.bind("8", functools.partial(score_selected_question_answersheet, "8"))
    self.window.bind("9", functools.partial(score_selected_question_answersheet, "9"))
    self.window.bind("<BackSpace>", functools.partial(score_selected_question_answersheet, "backspace"))
    self.window.bind("<Control-q>", functools.partial(toggle_booleanVar_checkbutton_show, "unscored"))
    self.window.bind("<Control-e>", functools.partial(toggle_booleanVar_checkbutton_show, "correct"))
    self.window.bind("<Control-f>", functools.partial(toggle_booleanVar_checkbutton_show, "partial"))
    self.window.bind("<Control-j>", functools.partial(toggle_booleanVar_checkbutton_show, "hold"))
    self.window.bind("<Control-o>", functools.partial(toggle_booleanVar_checkbutton_show, "incorrect"))

  @sub_window_loop
  def export(self):
    def export_list_xlsx():
      ### この関数いろいろダメです。信用しないで下さい。
      def set_style(table_cells: list[list[openpyxl.cell.cell.Cell]], *, internal_border=True, left_side_thin=True, right_side_thin=True) -> None:
        for index_rows, rows in enumerate(table_cells):
          for index_column, cell in enumerate(rows):
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal="center", vertical="center")
            cell.font = openpyxl.styles.Font(size=11, name="Meiryo UI")
            if internal_border: # 内側あり
              if index_column == 0 and left_side_thin: 
                cell.border = openpyxl.styles.borders.Border(
                  top=openpyxl.styles.Side(style="thin", color="000000"),
                  bottom=openpyxl.styles.Side(style="thin", color="000000"),
                  left=openpyxl.styles.Side(style="thin", color="000000"),
                  right=openpyxl.styles.Side(style="hair", color="000000")
                )
              elif index_column == len(rows) - 1 and right_side_thin:
                  cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.Side(style="thin", color="000000"),
                    bottom=openpyxl.styles.Side(style="thin", color="000000"),
                    left=openpyxl.styles.Side(style="hair", color="000000"),
                    right=openpyxl.styles.Side(style="thin", color="000000")
                  )
              else: 
                cell.border = openpyxl.styles.borders.Border(
                  top=openpyxl.styles.Side(style="thin", color="000000"),
                  bottom=openpyxl.styles.Side(style="thin", color="000000"),
                  left=openpyxl.styles.Side(style="hair", color="000000"),
                  right=openpyxl.styles.Side(style="hair", color="000000")
                )
            else: # 内側なし
              if index_rows == 0: # 最上行
                if len(rows) == 1:
                  if left_side_thin: 
                    if right_side_thin:
                      cell.border = openpyxl.styles.borders.Border(
                        top=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="thin", color="000000"),
                        right=openpyxl.styles.Side(style="thin", color="000000")
                      )
                    else:
                      cell.border = openpyxl.styles.borders.Border(
                        top=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="thin", color="000000"),
                        right=openpyxl.styles.Side(style="hair", color="000000")
                      )
                  else:
                    if right_side_thin:
                      cell.border = openpyxl.styles.borders.Border(
                        top=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="hair", color="000000"),
                        right=openpyxl.styles.Side(style="thin", color="000000")
                      )
                    else:
                      cell.border = openpyxl.styles.borders.Border(
                        top=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="hair", color="000000"),
                        right=openpyxl.styles.Side(style="hair", color="000000")
                      )
                elif index_column == 0:
                  if left_side_thin: 
                    cell.border = openpyxl.styles.borders.Border(
                      top=openpyxl.styles.Side(style="thin", color="000000"),
                      left=openpyxl.styles.Side(style="thin", color="000000")
                    )
                  else:
                    cell.border = openpyxl.styles.borders.Border(
                      top=openpyxl.styles.Side(style="thin", color="000000"),
                      left=openpyxl.styles.Side(style="hair", color="000000")
                    )
                elif index_column == len(rows) - 1:
                  if right_side_thin: 
                    cell.border = openpyxl.styles.borders.Border(
                      top=openpyxl.styles.Side(style="thin", color="000000"),
                      right=openpyxl.styles.Side(style="thin", color="000000")
                    )
                  else:
                    cell.border = openpyxl.styles.borders.Border(
                      top=openpyxl.styles.Side(style="thin", color="000000"),
                      right=openpyxl.styles.Side(style="hair", color="000000")
                    )
                else:
                  cell.border = openpyxl.styles.borders.Border(
                    top=openpyxl.styles.Side(style="thin", color="000000")
                  )
              elif index_rows == len(table_cells) - 1: # 最下行
                if len(rows) == 1:
                  if left_side_thin: 
                    if right_side_thin:
                      cell.border = openpyxl.styles.borders.Border(
                        bottom=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="thin", color="000000"),
                        right=openpyxl.styles.Side(style="thin", color="000000")
                      )
                    else:
                      cell.border = openpyxl.styles.borders.Border(
                        bottom=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="thin", color="000000"),
                        right=openpyxl.styles.Side(style="hair", color="000000")
                      )
                  else:
                    if right_side_thin:
                      cell.border = openpyxl.styles.borders.Border(
                        bottom=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="hair", color="000000"),
                        right=openpyxl.styles.Side(style="thin", color="000000")
                      )
                    else:
                      cell.border = openpyxl.styles.borders.Border(
                        bottom=openpyxl.styles.Side(style="thin", color="000000"),
                        left=openpyxl.styles.Side(style="hair", color="000000"),
                        right=openpyxl.styles.Side(style="hair", color="000000")
                      )
                elif index_column == 0:
                  if left_side_thin: 
                    cell.border = openpyxl.styles.borders.Border(
                      bottom=openpyxl.styles.Side(style="thin", color="000000"),
                      left=openpyxl.styles.Side(style="thin", color="000000")
                    )
                  else:
                    cell.border = openpyxl.styles.borders.Border(
                      bottom=openpyxl.styles.Side(style="thin", color="000000"),
                      left=openpyxl.styles.Side(style="hair", color="000000")
                    )
                elif index_column == len(rows) - 1:
                  if right_side_thin: 
                    cell.border = openpyxl.styles.borders.Border(
                      bottom=openpyxl.styles.Side(style="thin", color="000000"),
                      right=openpyxl.styles.Side(style="thin", color="000000")
                    )
                  else:
                    cell.border = openpyxl.styles.borders.Border(
                      bottom=openpyxl.styles.Side(style="thin", color="000000"),
                      right=openpyxl.styles.Side(style="hair", color="000000")
                    )
                else:
                  cell.border = openpyxl.styles.borders.Border(
                    bottom=openpyxl.styles.Side(style="thin", color="000000")
                  )
              else: # 中行
                if len(rows) == 1:
                  if left_side_thin: 
                    if right_side_thin:
                      cell.border = openpyxl.styles.borders.Border(
                        left=openpyxl.styles.Side(style="thin", color="000000"),
                        right=openpyxl.styles.Side(style="thin", color="000000")
                      )
                    else:
                      cell.border = openpyxl.styles.borders.Border(
                        left=openpyxl.styles.Side(style="thin", color="000000"),
                        right=openpyxl.styles.Side(style="hair", color="000000")
                      )
                  else:
                    if right_side_thin:
                      cell.border = openpyxl.styles.borders.Border(
                        left=openpyxl.styles.Side(style="hair", color="000000"),
                        right=openpyxl.styles.Side(style="thin", color="000000")
                      )
                    else:
                      cell.border = openpyxl.styles.borders.Border(
                        left=openpyxl.styles.Side(style="hair", color="000000"),
                        right=openpyxl.styles.Side(style="hair", color="000000")
                      )
                elif index_column == 0:
                  if left_side_thin: 
                    cell.border = openpyxl.styles.borders.Border(
                      left=openpyxl.styles.Side(style="thin", color="000000")
                    )
                  else:
                    cell.border = openpyxl.styles.borders.Border(
                      left=openpyxl.styles.Side(style="hair", color="000000")
                    )
                elif index_column == len(rows) - 1:
                  if right_side_thin: 
                    cell.border = openpyxl.styles.borders.Border(
                      right=openpyxl.styles.Side(style="thin", color="000000")
                    )
                  else:
                    cell.border = openpyxl.styles.borders.Border(
                      right=openpyxl.styles.Side(style="hair", color="000000")
                    )
                else:
                  cell.border = openpyxl.styles.borders.Border()
      
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
      path_dir = dict_project["path_dir"]
      with open(path_dir + "/.temp_saiten/answer_area.json", "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)
      with open(path_dir + "/.temp_saiten/load_picture.json", "r", encoding="utf-8") as f:
        dict_load_picture = json.load(f)
      with open(path_dir + "/.temp_saiten/meibo.json", "r", encoding="utf-8") as f:
        list_meibo = json.load(f)
      
      workbook_result_scoring = openpyxl.Workbook()
      workbook_result_scoring.remove(workbook_result_scoring["Sheet"])
      workbook_result_scoring.create_sheet(title="点数一覧")
      workbook_result_scoring.create_sheet(title="正誤一覧")
      list_daimon = list(set([question["daimon"] for question in dict_answer_area["questions"] if question["type"] == "設問"]))
      list_daimon.sort()
      list_name_gakunen = list(set([meibo["学年"] for meibo in list_meibo]))
      list_tuple_gakkyuu = list(set([(meibo["学年"], meibo["学級"]) for meibo in list_meibo]))
      list_tuple_gakkyuu.sort(key=lambda x:(x[0], x[1]))
      # workbook_result_scoring["点数一覧"].views.SheetView(showGridLines=False) # 目盛線を非表示
      # 答案用紙ごとのスコアのリスト
      list_list_score = [[dict_answer_area["questions"][index_question]["score"][index_answersheet] for index_question in range(len(dict_answer_area["questions"])) if dict_answer_area["questions"][index_question]["type"] == "設問"] for index_answersheet in range(len(list_meibo))]
      list_tuple_question = [(question["daimon"], question["shomon"], question["shimon"], question["haiten"]) for question in dict_answer_area["questions"] if question["type"] == "設問"]
      list_list_score_point = []
      list_list_score_status = []
      for index_list_score, list_score in enumerate(list_list_score):
        list_list_score_point.append([])
        list_list_score_status.append([])
        for index_score, score in enumerate(list_score):
          if score["status"] == "unscored":
            list_list_score_point[-1].append("")
            list_list_score_status[-1].append(f"-")
          elif score["status"] == "correct":
            list_list_score_point[-1].append(list_tuple_question[index_score][3])
            list_list_score_status[-1].append(f"○")
          elif score["status"] == "partial":
            list_list_score_point[-1].append(score["point"])
            list_list_score_status[-1].append(f"△{score['point']}")
          elif score["status"] == "hold":
            list_list_score_point[-1].append(score["point"])
            list_list_score_status[-1].append(f"？{list_tuple_question[index_score][3]}")
          elif score["status"] == "incorrect":
            list_list_score_point[-1].append(0)
            list_list_score_status[-1].append(f"×")
      tuple_rowrange_gakunen = (7, 6 + len(list_name_gakunen))
      tuple_rowrange_gakkyuu = (7 + len(list_name_gakunen), 6 + len(list_name_gakunen) + len(list_tuple_gakkyuu))
      tuple_rowrange_meibo = (7 + len(list_name_gakunen) + len(list_tuple_gakkyuu), 6 + len(list_name_gakunen) + len(list_tuple_gakkyuu) + len(list_meibo))
      tuple_columnrange_goukei = (7, 6 + 1)
      tuple_columnrange_shoukei = (7 + 1, 6 + 1 + len(list_daimon))
      tuple_columnrange_question = (7 + 1 + len(list_daimon), 6 + 1 + len(list_daimon) + len([question for question in dict_answer_area["questions"] if question["type"] == "設問"]))

      for sheet in [workbook_result_scoring["点数一覧"], workbook_result_scoring["正誤一覧"]]:

        # 表全体の書式設定 (中央揃え / フォント)
        set_style(sheet[f"B2:{openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1] + 4)}{tuple_rowrange_meibo[1]}"])
        sheet.row_dimensions[1].height = 5 * 3 / 4
        sheet.column_dimensions["A"].width = 5 / 8
        sheet.column_dimensions["B"].width = 60 / 8
        sheet.column_dimensions["C"].width = 60 / 8
        sheet.column_dimensions["D"].width = 60 / 8
        sheet.column_dimensions["E"].width = 80 / 8
        sheet.column_dimensions["F"].width = 80 / 8
        sheet.freeze_panes = "G7"

        # row: 2-6 
        ### column B-F
        sheet["B3"].value = dict_project["name"]
        sheet["B4"].value = f"採点結果 - {sheet.title}"
        set_style(sheet["B2:E5"], internal_border=False, right_side_thin=False)
        for rows in sheet["B2:E5"]:
          for cell in rows:
            cell.alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")
        sheet["F2"].value = "大問"
        sheet["F3"].value = "小問"
        sheet["F4"].value = "枝問"
        sheet["F5"].value = "配点"

        sheet["B6"].value = "学年"
        sheet["C6"].value = "学級"
        sheet["D6"].value = "出席番号"
        sheet["E6"].value = "生徒番号"
        sheet["F6"].value = "氏名"

        ### column: 合計得点
        sheet["G2"].value = "合"
        sheet["G3"].value = "計"
        if sheet.title == "点数一覧":
          sheet["G4"].value = "得"
          sheet["G5"].value = "点"
        else:
          sheet["G4"].value = "設問"
          sheet["G5"].value = "正答数"
        set_style(sheet[f"G2:G5"], internal_border=False, left_side_thin=False, right_side_thin=False)
        ### column: 各大問ごとの小計点
        for index_daimon, daimon in enumerate(list_daimon):
          sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)].width = 50 / 8
          sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=2).value = daimon
          if sheet.title == "点数一覧":
            sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=3).value = "小"
            sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=4).value = "計"
            sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=5).value = "点"
          else:
            sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=3).value = "小計"
            sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=4).value = "設問"
            sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=5).value = "正答数"
          set_style(sheet[f"{openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)}2:{openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)}5"], internal_border=False, left_side_thin=False, right_side_thin=False)
        ### column: 各設問の 大問 / 小問 / 枝問 / 配点
        for index_tuple_question, tuple_question in enumerate(list_tuple_question):
          sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0] + index_tuple_question)].width = 40 / 8
          sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=2).value = tuple_question[0]
          sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=3).value = tuple_question[1]
          sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=4).value = tuple_question[2]
          sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=5).value = tuple_question[3]

        ### 順位, 生徒番号, 氏名
        sheet.cell(column=tuple_columnrange_question[1] + 1, row=2).value = "学"
        sheet.cell(column=tuple_columnrange_question[1] + 1, row=3).value = "年"
        sheet.cell(column=tuple_columnrange_question[1] + 1, row=4).value = "順"
        sheet.cell(column=tuple_columnrange_question[1] + 1, row=5).value = "位"
        sheet.cell(column=tuple_columnrange_question[1] + 2, row=2).value = "学"
        sheet.cell(column=tuple_columnrange_question[1] + 2, row=3).value = "級"
        sheet.cell(column=tuple_columnrange_question[1] + 2, row=4).value = "順"
        sheet.cell(column=tuple_columnrange_question[1] + 2, row=5).value = "位"
        sheet.cell(column=tuple_columnrange_question[1] + 3, row=6).value = "生徒番号"
        sheet.cell(column=tuple_columnrange_question[1] + 4, row=6).value = "氏名"
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1] + 1)].width = 30 / 8
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1] + 2)].width = 30 / 8
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1] + 3)].width = 80 / 8
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1] + 4)].width = 80 / 8
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1] + 5)].width = 5 / 8

        # row: 学年平均点 / 学年正答率
        for index_name_gakunen, name_gakunen in enumerate(list_name_gakunen):
          sheet.cell(row=tuple_rowrange_gakunen[0] + index_name_gakunen, column=2).value = name_gakunen
          for index_column in [2, 3, 4, 5, 6]:
            sheet.cell(row=tuple_rowrange_gakunen[0] + index_name_gakunen, column=index_column).alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")
          if sheet.title == "点数一覧":
            sheet.cell(row=tuple_rowrange_gakunen[0] + index_name_gakunen, column=3).value = "学年平均点"
            for index_column in [index_column + 7 for index_column in range(1 + len(list_daimon) + len(list_tuple_question))]:
              sheet.cell(column=index_column, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = f"=AVERAGEIFS({openpyxl.utils.cell.get_column_letter(index_column)}${tuple_rowrange_meibo[0]}:{openpyxl.utils.cell.get_column_letter(index_column)}${tuple_rowrange_meibo[1]}, $B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_gakunen[0] + index_name_gakunen})"
              sheet.cell(column=index_column, row=tuple_rowrange_gakunen[0] + index_name_gakunen).number_format = "0.0"
          else:
            sheet.cell(row=tuple_rowrange_gakunen[0] + index_name_gakunen, column=3).value = "学年正答率"
            sheet.cell(column=7, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = f"=AVERAGE(${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}{tuple_rowrange_gakunen[0] + index_name_gakunen}:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}{tuple_rowrange_gakunen[0] + index_name_gakunen})"
            sheet.cell(column=7, row=tuple_rowrange_gakunen[0] + index_name_gakunen).number_format = "[=1]1;.000"
            for index_daimon, daimon in enumerate(list_daimon):
              sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = f"=AVERAGEIFS(${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}{tuple_rowrange_gakunen[0] + index_name_gakunen}:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}{tuple_rowrange_gakunen[0] + index_name_gakunen}, ${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}$2:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}$2, {openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)}$2)"
              sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=tuple_rowrange_gakunen[0] + index_name_gakunen).number_format = "[=1]1;.000"
            for index_tuple_question, tuple_question in enumerate(list_tuple_question):
              sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = f"=COUNTIFS({openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0] + index_tuple_question)}${tuple_rowrange_meibo[0]}:{openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0] + index_tuple_question)}${tuple_rowrange_meibo[1]}, \"○\", $B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_gakunen[0] + index_name_gakunen})/COUNTIFS($B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_gakunen[0] + index_name_gakunen})"
              sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=tuple_rowrange_gakunen[0] + index_name_gakunen).number_format = "[=1]1;.000"
          sheet.cell(column=tuple_columnrange_question[1] + 1, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = "-"
          sheet.cell(column=tuple_columnrange_question[1] + 2, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = "-"
          sheet.cell(column=tuple_columnrange_question[1] + 3, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = "-"
          sheet.cell(column=tuple_columnrange_question[1] + 4, row=tuple_rowrange_gakunen[0] + index_name_gakunen).value = "-"

        # row: 学級平均点 / 学級平均正答数
        for index_tuple_gakkyuu, tuple_gakkyuu in enumerate(list_tuple_gakkyuu):
          sheet.cell(row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu, column=2).value = tuple_gakkyuu[0]
          sheet.cell(row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu, column=3).value = tuple_gakkyuu[1]
          if sheet.title == "点数一覧":
            sheet.cell(row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu, column=4).value = "学級平均点"
          else:
            sheet.cell(row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu, column=4).value = "学級正答率"
          for index_column in [3, 4, 5, 6]:
            sheet.cell(row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu, column=index_column).alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")
          if sheet.title == "点数一覧":
            for index_column in [index_column + 7 for index_column in range(1 + len(list_daimon) + len(list_tuple_question))]:
              sheet.cell(column=index_column, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = f"=AVERAGEIFS(${openpyxl.utils.cell.get_column_letter(index_column)}${tuple_rowrange_meibo[0]}:${openpyxl.utils.cell.get_column_letter(index_column)}${tuple_rowrange_meibo[1]}, $B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu}, $C${tuple_rowrange_meibo[0]}:$C${tuple_rowrange_meibo[1]}, $C{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu})"
              sheet.cell(column=index_column, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).number_format = "0.0"
          else:
            sheet.cell(column=7, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = f"=AVERAGE(${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu}:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu})"
            sheet.cell(column=7, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).number_format = "[=1]1;.000"
            for index_daimon, daimon in enumerate(list_daimon):
              sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = f"=AVERAGEIFS(${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu}:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu}, ${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}$2:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}$2, {openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)}$2)"
              sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).number_format = "[=1]1;.000"
            for index_tuple_question, tuple_question in enumerate(list_tuple_question):
              sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = f"=COUNTIFS({openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0] + index_tuple_question)}${tuple_rowrange_meibo[0]}:{openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0] + index_tuple_question)}${tuple_rowrange_meibo[1]}, \"○\", $B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu}, $C${tuple_rowrange_meibo[0]}:$C${tuple_rowrange_meibo[1]}, $C{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu})/COUNTIFS($B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu}, $C${tuple_rowrange_meibo[0]}:$C${tuple_rowrange_meibo[1]}, $C{tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu})"
              sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).number_format = "[=1]1;.000"
          sheet.cell(column=tuple_columnrange_question[1] + 1, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = "-"
          sheet.cell(column=tuple_columnrange_question[1] + 2, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = "-"
          sheet.cell(column=tuple_columnrange_question[1] + 3, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = "-"
          sheet.cell(column=tuple_columnrange_question[1] + 4, row=tuple_rowrange_gakkyuu[0] + index_tuple_gakkyuu).value = "-"

        # row: 名簿
        for index_meibo, meibo in enumerate(list_meibo):
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=2).value = meibo["学年"]
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=3).value = meibo["学級"]
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=4).value = meibo["出席番号"]
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=5).value = meibo["生徒番号"]
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=6).value = meibo["氏名"]
          if sheet.title == "点数一覧":
            ### column: 合計得点
            sheet.cell(column=tuple_columnrange_goukei[0], row=tuple_rowrange_meibo[0] + index_meibo).value = f"=SUM({openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}${tuple_rowrange_meibo[0] + index_meibo}:{openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}${tuple_rowrange_meibo[0] + index_meibo})"
            ### column: 各大問ごとの小計点
            for index_daimon, daimon in enumerate(list_daimon):
              sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=tuple_rowrange_meibo[0] + index_meibo).value = f"=SUMIFS(${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}{tuple_rowrange_meibo[0] + index_meibo}:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}{tuple_rowrange_meibo[0] + index_meibo}, ${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}$2:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}$2, {openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)}$2)"
            ### 各設問
            for index_tuple_question in range(len(list_tuple_question)):
              sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=tuple_rowrange_meibo[0] + index_meibo).value = list_list_score_point[index_meibo][index_tuple_question]
          else:
            ### column: 合計正答設問数
            sheet.cell(column=tuple_columnrange_goukei[0], row=tuple_rowrange_meibo[0] + index_meibo).value = f"=COUNTIFS({openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}${tuple_rowrange_meibo[0] + index_meibo}:{openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}${tuple_rowrange_meibo[0] + index_meibo}, \"○\")"
            ### column: 各大問ごとの小計正答設問数
            for index_daimon, daimon in enumerate(list_daimon):
              sheet.cell(column=tuple_columnrange_shoukei[0] + index_daimon, row=tuple_rowrange_meibo[0] + index_meibo).value = f"=COUNTIFS(${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}{tuple_rowrange_meibo[0] + index_meibo}:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}{tuple_rowrange_meibo[0] + index_meibo}, \"○\", ${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[0])}$2:${openpyxl.utils.cell.get_column_letter(tuple_columnrange_question[1])}$2, {openpyxl.utils.cell.get_column_letter(tuple_columnrange_shoukei[0] + index_daimon)}$2)"
            ### 設問
            for index_tuple_question in range(len(list_tuple_question)):
              sheet.cell(column=tuple_columnrange_question[0] + index_tuple_question, row=tuple_rowrange_meibo[0] + index_meibo).value = list_list_score_status[index_meibo][index_tuple_question]
          sheet.cell(column=tuple_columnrange_question[1] + 1, row=tuple_rowrange_meibo[0] + index_meibo).value = f"=COUNTIFS($B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_meibo[0] + index_meibo}, $G${tuple_rowrange_meibo[0]}:$G${tuple_rowrange_meibo[1]}, \">\"&$G{tuple_rowrange_meibo[0] + index_meibo}) + 1"
          sheet.cell(column=tuple_columnrange_question[1] + 2, row=tuple_rowrange_meibo[0] + index_meibo).value = f"=COUNTIFS($B${tuple_rowrange_meibo[0]}:$B${tuple_rowrange_meibo[1]}, $B{tuple_rowrange_meibo[0] + index_meibo}, $C${tuple_rowrange_meibo[0]}:$C${tuple_rowrange_meibo[1]}, $C{tuple_rowrange_meibo[0] + index_meibo}, $G${tuple_rowrange_meibo[0]}:$G${tuple_rowrange_meibo[1]}, \">\"&$G{tuple_rowrange_meibo[0] + index_meibo}) + 1"
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=tuple_columnrange_question[1] + 3).value = meibo["生徒番号"]
          sheet.cell(row=tuple_rowrange_meibo[0] + index_meibo, column=tuple_columnrange_question[1] + 4).value = meibo["氏名"]

      path_workbook_result_scoring = tkinter.filedialog.asksaveasfile(
        parent=self.window,
        title = "採点データを名前を付けて保存",
        filetypes=[("Excel スプレッドシート", ".xlsx")],
        defaultextension="xlsx"
      )
      try:
        workbook_result_scoring.save(path_workbook_result_scoring.name)
      except PermissionError:
        tkinter.messagebox.showerror(
          "ファイルを保存できません",
          "ファイルを保存できませんでした. \n"
          + "ファイルを開いていませんか？\n"
          + "Excel を終了して, もう一度お試し下さい. "
        )

    def preview_export_picture():
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
      path_dir = dict_project["path_dir"]
      path_json_answer_area = dict_project["path_dir"] + "/.temp_saiten/answer_area.json"
      path_file_model_answer = dict_project["path_dir"] + "/.temp_saiten/model_answer/model_answer.png"
      path_dir_of_answers = dict_project["path_dir"] + "/.temp_saiten/answer"
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)

      canvas.delete("saiten")

      size = dict_project["export"]["symbol"]["size"]
      self.dict_image_scoring_symbol = {
        "unscored": PIL.Image.open(os.path.dirname(__file__) + "/assets/unscored.png"),
        "correct": PIL.Image.open(os.path.dirname(__file__) + "/assets/correct.png"),
        "partial": PIL.Image.open(os.path.dirname(__file__) + "/assets/partial.png"),
        "hold": PIL.Image.open(os.path.dirname(__file__) + "/assets/hold.png"),
        "incorrect": PIL.Image.open(os.path.dirname(__file__) + "/assets/incorrect.png"),
        "tranceparent_unscored": PIL.Image.open(os.path.dirname(__file__) + "/assets/tranceparent_unscored.png"),
        "tranceparent_correct": PIL.Image.open(os.path.dirname(__file__) + "/assets/tranceparent_correct.png"),
        "tranceparent_partial": PIL.Image.open(os.path.dirname(__file__) + "/assets/tranceparent_partial.png"),
        "tranceparent_hold": PIL.Image.open(os.path.dirname(__file__) + "/assets/tranceparent_hold.png"),
        "tranceparent_incorrect": PIL.Image.open(os.path.dirname(__file__) + "/assets/tranceparent_incorrect.png")
      }
      self.dict_image_scoring_symbol_resized = {
        "unscored": self.dict_image_scoring_symbol["unscored"].resize((size, size)),
        "correct": self.dict_image_scoring_symbol["correct"].resize((size, size)),
        "partial": self.dict_image_scoring_symbol["partial"].resize((size, size)),
        "hold": self.dict_image_scoring_symbol["hold"].resize((size, size)),
        "incorrect": self.dict_image_scoring_symbol["incorrect"].resize((size, size)),
        "tranceparent_unscored": self.dict_image_scoring_symbol["tranceparent_unscored"].resize((size, size)),
        "tranceparent_correct": self.dict_image_scoring_symbol["tranceparent_correct"].resize((size, size)),
        "tranceparent_partial": self.dict_image_scoring_symbol["tranceparent_partial"].resize((size, size)),
        "tranceparent_hold": self.dict_image_scoring_symbol["tranceparent_hold"].resize((size, size)),
        "tranceparent_incorrect": self.dict_image_scoring_symbol["tranceparent_incorrect"].resize((size, size))
      }
      self.dict_imagetk_scoring_symbol = {
        "unscored": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["unscored"]),
        "correct": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["correct"]),
        "partial": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["partial"]),
        "hold": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["hold"]),
        "incorrect": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["incorrect"]),
        "tranceparent_unscored": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["tranceparent_unscored"]),
        "tranceparent_correct": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["tranceparent_correct"]),
        "tranceparent_partial": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["tranceparent_partial"]),
        "tranceparent_hold": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["tranceparent_hold"]),
        "tranceparent_incorrect": PIL.ImageTk.PhotoImage(image=self.dict_image_scoring_symbol_resized["tranceparent_incorrect"])
      }
      index_setsumon = 0
      booleanvar_unscored_symbol.set(value=dict_project["export"]["symbol"]["unscored"])
      booleanvar_correct_symbol.set(value=dict_project["export"]["symbol"]["correct"])
      booleanvar_partial_symbol.set(value=dict_project["export"]["symbol"]["partial"])
      booleanvar_hold_symbol.set(value=dict_project["export"]["symbol"]["hold"])
      booleanvar_incorrect_symbol.set(value=dict_project["export"]["symbol"]["incorrect"])
      booleanvar_unscored_point.set(value=dict_project["export"]["point"]["unscored"])
      booleanvar_correct_point.set(value=dict_project["export"]["point"]["correct"])
      booleanvar_partial_point.set(value=dict_project["export"]["point"]["partial"])
      booleanvar_hold_point.set(value=dict_project["export"]["point"]["hold"])
      booleanvar_incorrect_point.set(value=dict_project["export"]["point"]["incorrect"])
      for question in dict_answer_area["questions"]:
        if question["type"] == "設問":
          index_setsumon += 1
          if dict_project["export"]["symbol"]["position"] == "nw":
            position_x = question["area"][0] + dict_project["export"]["symbol"]["x"]
            position_y = question["area"][1] + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "n":
            position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["symbol"]["x"]
            position_y = question["area"][1] + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "ne":
            position_x = question["area"][2] + dict_project["export"]["symbol"]["x"]
            position_y = question["area"][1] + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "w":
            position_x = question["area"][0] + dict_project["export"]["symbol"]["x"]
            position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "c":
            position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["symbol"]["x"]
            position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "e":
            position_x = question["area"][2] + dict_project["export"]["symbol"]["x"]
            position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "sw":
            position_x = question["area"][0] + dict_project["export"]["symbol"]["x"]
            position_y = question["area"][3] + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "s":
            position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["symbol"]["x"]
            position_y = question["area"][3] + dict_project["export"]["symbol"]["y"]
          elif dict_project["export"]["symbol"]["position"] == "se":
            position_x = question["area"][2] + dict_project["export"]["symbol"]["x"]
            position_y = question["area"][3] + dict_project["export"]["symbol"]["y"]
          if index_setsumon % 5 == 0 and booleanvar_unscored_symbol.get():
            canvas.create_image(position_x, position_y, anchor="center", image=self.dict_imagetk_scoring_symbol["tranceparent_unscored"], tags="saiten")
          elif index_setsumon % 5 == 1 and booleanvar_correct_symbol.get():
            canvas.create_image(position_x, position_y, anchor="center", image=self.dict_imagetk_scoring_symbol["tranceparent_correct"], tags="saiten")
          elif index_setsumon % 5 == 2 and booleanvar_partial_symbol.get():
            canvas.create_image(position_x, position_y, anchor="center", image=self.dict_imagetk_scoring_symbol["tranceparent_partial"], tags="saiten")
          elif index_setsumon % 5 == 3 and booleanvar_hold_symbol.get():
            canvas.create_image(position_x, position_y, anchor="center", image=self.dict_imagetk_scoring_symbol["tranceparent_hold"], tags="saiten")
          elif index_setsumon % 5 == 4 and booleanvar_incorrect_symbol.get():
            canvas.create_image(position_x, position_y, anchor="center", image=self.dict_imagetk_scoring_symbol["tranceparent_incorrect"], tags="saiten")
      index_setsumon = 0
      for question in dict_answer_area["questions"]:
        if question["type"] == "設問":
          index_setsumon += 1
          if dict_project["export"]["point"]["position"] == "nw":
            position_x = question["area"][0] + dict_project["export"]["point"]["x"]
            position_y = question["area"][1] + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "n":
            position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["point"]["x"]
            position_y = question["area"][1] + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "ne":
            position_x = question["area"][2] + dict_project["export"]["point"]["x"]
            position_y = question["area"][1] + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "w":
            position_x = question["area"][0] + dict_project["export"]["point"]["x"]
            position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "c":
            position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["point"]["x"]
            position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "e":
            position_x = question["area"][2] + dict_project["export"]["point"]["x"]
            position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "sw":
            position_x = question["area"][0] + dict_project["export"]["point"]["x"]
            position_y = question["area"][3] + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "s":
            position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["point"]["x"]
            position_y = question["area"][3] + dict_project["export"]["point"]["y"]
          elif dict_project["export"]["point"]["position"] == "se":
            position_x = question["area"][2] + dict_project["export"]["point"]["x"]
            position_y = question["area"][3] + dict_project["export"]["point"]["y"]
          if index_setsumon % 5 == 0 and booleanvar_unscored_point.get():
            canvas.create_text(position_x, position_y, text=0, fill="red", font=("Meiryo UI", dict_project["export"]["point"]["size"], "roman"), tags="saiten")
          elif index_setsumon % 5 == 1 and booleanvar_correct_point.get():
            canvas.create_text(position_x, position_y, text=question["haiten"], fill="red", font=("Meiryo UI", dict_project["export"]["point"]["size"], "roman"), tags="saiten")
          elif index_setsumon % 5 == 2 and booleanvar_partial_point.get():
            canvas.create_text(position_x, position_y, text=question["haiten"] // 2, fill="red", font=("Meiryo UI", dict_project["export"]["point"]["size"], "roman"), tags="saiten")
          elif index_setsumon % 5 == 3 and booleanvar_hold_point.get():
            canvas.create_text(position_x, position_y, text=question["haiten"] // 2, fill="red", font=("Meiryo UI", dict_project["export"]["point"]["size"], "roman"), tags="saiten")
          elif index_setsumon % 5 == 4 and booleanvar_incorrect_point.get():
            canvas.create_text(position_x, position_y, text=0, fill="red", font=("Meiryo UI", dict_project["export"]["point"]["size"], "roman"), tags="saiten")

    def set_position(symbol_or_point, key_property, position, *args):
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      if key_property in ["position"]:
        dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"][symbol_or_point][key_property] = position
        with open("config.json", "w", encoding="utf-8") as f:
          json.dump(dict_config, f, indent=2)
        preview_export_picture()
      elif key_property in ["unscored", "correct", "partial", "hold", "incorrect"]:
        dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"][symbol_or_point][key_property] = not dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"][symbol_or_point][key_property]
        with open("config.json", "w", encoding="utf-8") as f:
          json.dump(dict_config, f, indent=2)
        preview_export_picture()
      elif key_property in ["x", "y", "size"]:
        if position == "":
          position = "0"
        if position in [str(i) for i in range(-10000, 10000)]:
          if dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"][symbol_or_point][key_property] == int(position):
            return True
          else:
            dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"][symbol_or_point][key_property] = int(position)
            with open("config.json", "w", encoding="utf-8") as f:
              json.dump(dict_config, f, indent=2)
            preview_export_picture()
            return True
        else:
          return False

    def set_position_ex1(*args):
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["position"] = "w"
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["x"] = 0
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["y"] = 0
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["size"] = 60
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["unscored"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["correct"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["partial"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["hold"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["incorrect"] = True
      entry_symbol_x.delete(0, tkinter.END)
      entry_symbol_y.delete(0, tkinter.END)
      entry_symbol_size.delete(0, tkinter.END)
      entry_symbol_x.insert(0, 0)
      entry_symbol_y.insert(0, 0)
      entry_symbol_size.insert(0, 60)
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["position"] = "w"
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["x"] = 0
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["y"] = 0
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["size"] = 15
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["unscored"] = False
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["correct"] = False
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["partial"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["hold"] = False
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["incorrect"] = False
      entry_point_x.delete(0, tkinter.END)
      entry_point_y.delete(0, tkinter.END)
      entry_point_size.delete(0, tkinter.END)
      entry_point_x.insert(0, 0)
      entry_point_y.insert(0, 0)
      entry_point_size.insert(0, 15)
      with open("config.json", "w", encoding="utf-8") as f:
        json.dump(dict_config, f, indent=2)
      preview_export_picture()

    def set_position_ex2(*args):
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["position"] = "c"
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["x"] = 0
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["y"] = 0
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["size"] = 60
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["unscored"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["correct"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["partial"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["hold"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["symbol"]["incorrect"] = True
      entry_symbol_x.delete(0, tkinter.END)
      entry_symbol_y.delete(0, tkinter.END)
      entry_symbol_size.delete(0, tkinter.END)
      entry_symbol_x.insert(0, 0)
      entry_symbol_y.insert(0, 0)
      entry_symbol_size.insert(0, 60)
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["position"] = "se"
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["x"] = -10
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["y"] = -10
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["size"] = 10
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["unscored"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["correct"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["partial"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["hold"] = True
      dict_config["projects"][dict_config["index_projects_in_listbox"]]["export"]["point"]["incorrect"] = True
      entry_point_x.delete(0, tkinter.END)
      entry_point_y.delete(0, tkinter.END)
      entry_point_size.delete(0, tkinter.END)
      entry_point_x.insert(0, -10)
      entry_point_y.insert(0, -10)
      entry_point_size.insert(0, 10)
      with open("config.json", "w", encoding="utf-8") as f:
        json.dump(dict_config, f, indent=2)
      preview_export_picture()

    def export_pdf():
      with open("config.json", "r", encoding="utf-8") as f:
        dict_config = json.load(f)
      dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
      path_dir = dict_project["path_dir"]
      path_json_answer_area = dict_project["path_dir"] + "/.temp_saiten/answer_area.json"
      path_json_meibo = dict_project["path_dir"] + "/.temp_saiten/meibo.json"
      path_file_model_answer = dict_project["path_dir"] + "/.temp_saiten/model_answer/model_answer.png"
      path_dir_of_answers = dict_project["path_dir"] + "/.temp_saiten/answer"
      with open(path_json_answer_area, "r", encoding="utf-8") as f:
        dict_answer_area = json.load(f)
      with open(path_json_meibo, "r", encoding="utf-8") as f:
        list_meibo = json.load(f)
      if not os.path.exists(f"{path_dir}/.temp_saiten/output"):
        os.mkdir(f"{path_dir}/.temp_saiten/output")
      
      self.list_image_answersheet = []
      self.dict_image_scoring_symbol_resized["tranceparent_unscored"] = self.dict_image_scoring_symbol_resized["tranceparent_unscored"].convert("RGBA")
      self.dict_image_scoring_symbol_resized["tranceparent_correct"] = self.dict_image_scoring_symbol_resized["tranceparent_correct"].convert("RGBA")
      self.dict_image_scoring_symbol_resized["tranceparent_partial"] = self.dict_image_scoring_symbol_resized["tranceparent_partial"].convert("RGBA")
      self.dict_image_scoring_symbol_resized["tranceparent_hold"] = self.dict_image_scoring_symbol_resized["tranceparent_hold"].convert("RGBA")
      self.dict_image_scoring_symbol_resized["tranceparent_incorrect"] = self.dict_image_scoring_symbol_resized["tranceparent_incorrect"].convert("RGBA")
      for index_meibo, meibo in enumerate(list_meibo):
        self.image_answersheet = PIL.Image.open(f"{path_dir_of_answers}/{index_meibo}.png").convert("RGBA")
        for question in dict_answer_area["questions"]:
          if question["type"] == "設問":
            if dict_project["export"]["symbol"]["position"] == "nw":
              position_x = question["area"][0] + dict_project["export"]["symbol"]["x"]
              position_y = question["area"][1] + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "n":
              position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["symbol"]["x"]
              position_y = question["area"][1] + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "ne":
              position_x = question["area"][2] + dict_project["export"]["symbol"]["x"]
              position_y = question["area"][1] + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "w":
              position_x = question["area"][0] + dict_project["export"]["symbol"]["x"]
              position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "c":
              position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["symbol"]["x"]
              position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "e":
              position_x = question["area"][2] + dict_project["export"]["symbol"]["x"]
              position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "sw":
              position_x = question["area"][0] + dict_project["export"]["symbol"]["x"]
              position_y = question["area"][3] + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "s":
              position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["symbol"]["x"]
              position_y = question["area"][3] + dict_project["export"]["symbol"]["y"]
            elif dict_project["export"]["symbol"]["position"] == "se":
              position_x = question["area"][2] + dict_project["export"]["symbol"]["x"]
              position_y = question["area"][3] + dict_project["export"]["symbol"]["y"]
            position_x -= dict_project["export"]["symbol"]["size"] // 2
            position_y -= dict_project["export"]["symbol"]["size"] // 2
            self.image_clear = PIL.Image.new("RGBA", self.image_answersheet.size, (255, 255, 255, 0))
            if question["score"][index_meibo]["status"] == "unscored" and booleanvar_unscored_symbol.get():
              self.image_clear.paste(self.dict_image_scoring_symbol_resized["tranceparent_unscored"], (position_x, position_y))
            elif question["score"][index_meibo]["status"] == "correct" and booleanvar_correct_symbol.get():
              self.image_clear.paste(self.dict_image_scoring_symbol_resized["tranceparent_correct"], (position_x, position_y))
            elif question["score"][index_meibo]["status"] == "partial" and booleanvar_partial_symbol.get():
              self.image_clear.paste(self.dict_image_scoring_symbol_resized["tranceparent_partial"], (position_x, position_y))
            elif question["score"][index_meibo]["status"] == "hold" and booleanvar_hold_symbol.get():
              self.image_clear.paste(self.dict_image_scoring_symbol_resized["tranceparent_hold"], (position_x, position_y))
            elif question["score"][index_meibo]["status"] == "incorrect" and booleanvar_incorrect_symbol.get():
              self.image_clear.paste(self.dict_image_scoring_symbol_resized["tranceparent_incorrect"], (position_x, position_y))
            self.image_answersheet = PIL.Image.alpha_composite(self.image_answersheet, self.image_clear)
        for question in dict_answer_area["questions"]:
          if question["type"] == "設問":
            if dict_project["export"]["point"]["position"] == "nw":
              position_x = question["area"][0] + dict_project["export"]["point"]["x"]
              position_y = question["area"][1] + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "n":
              position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["point"]["x"]
              position_y = question["area"][1] + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "ne":
              position_x = question["area"][2] + dict_project["export"]["point"]["x"]
              position_y = question["area"][1] + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "w":
              position_x = question["area"][0] + dict_project["export"]["point"]["x"]
              position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "c":
              position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["point"]["x"]
              position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "e":
              position_x = question["area"][2] + dict_project["export"]["point"]["x"]
              position_y = (question["area"][1] + question["area"][3]) // 2 + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "sw":
              position_x = question["area"][0] + dict_project["export"]["point"]["x"]
              position_y = question["area"][3] + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "s":
              position_x = (question["area"][0] + question["area"][2]) // 2 + dict_project["export"]["point"]["x"]
              position_y = question["area"][3] + dict_project["export"]["point"]["y"]
            elif dict_project["export"]["point"]["position"] == "se":
              position_x = question["area"][2] + dict_project["export"]["point"]["x"]
              position_y = question["area"][3] + dict_project["export"]["point"]["y"]
            position_x -= dict_project["export"]["point"]["size"] // 2
            position_y -= dict_project["export"]["point"]["size"] // 2
            if question["score"][index_meibo]["status"] == "unscored" and booleanvar_unscored_point.get():
              PIL.ImageDraw.Draw(self.image_answersheet).text((position_x, position_y), str(0), fill="red", font=PIL.ImageFont.truetype("meiryo.ttc", size=dict_project["export"]["point"]["size"]))
            elif question["score"][index_meibo]["status"] == "correct" and booleanvar_correct_point.get():
              PIL.ImageDraw.Draw(self.image_answersheet).text((position_x, position_y), str(question["haiten"]), fill="red", font=PIL.ImageFont.truetype("meiryo.ttc", size=dict_project["export"]["point"]["size"]))
            elif question["score"][index_meibo]["status"] == "partial" and booleanvar_partial_point.get():
              PIL.ImageDraw.Draw(self.image_answersheet).text((position_x, position_y), str(question["score"][index_meibo]["point"]), fill="red", font=PIL.ImageFont.truetype("meiryo.ttc", size=dict_project["export"]["point"]["size"]))
            elif question["score"][index_meibo]["status"] == "hold" and booleanvar_hold_point.get():
              PIL.ImageDraw.Draw(self.image_answersheet).text((position_x, position_y), str(question["score"][index_meibo]["point"]), fill="red", font=PIL.ImageFont.truetype("meiryo.ttc", size=dict_project["export"]["point"]["size"]))
            elif question["score"][index_meibo]["status"] == "incorrect" and booleanvar_incorrect_point.get():
              PIL.ImageDraw.Draw(self.image_answersheet).text((position_x, position_y), str(0), fill="red", font=PIL.ImageFont.truetype("meiryo.ttc", size=dict_project["export"]["point"]["size"]))
        self.image_answersheet.save(f"{path_dir}/.temp_saiten/output/{index_meibo}.png")
      
      path_pdf = tkinter.filedialog.asksaveasfile(
        parent=self.window,
        title="採点済答案画像の出力",
        filetypes=[("PDF ドキュメント", ".pdf")],
        defaultextension="pdf"
      )
      if path_pdf is not None:
        with open(path_pdf.name, "wb") as f:
          f.write(img2pdf.convert([PIL.Image.open(f"{path_dir}/.temp_saiten/output/{index_meibo}.png").filename for index_meibo, meibo in enumerate(list_meibo)]))

    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
    path_dir = dict_project["path_dir"]
    path_json_answer_area = dict_project["path_dir"] + "/.temp_saiten/answer_area.json"
    path_file_model_answer = dict_project["path_dir"] + "/.temp_saiten/model_answer/model_answer.png"
    path_dir_of_answers = dict_project["path_dir"] + "/.temp_saiten/answer"
    with open(path_json_answer_area, "r", encoding="utf-8") as f:
      dict_answer_area = json.load(f)
      
    self.window.title("書き出し")
    self.canvas_draw_rectangle = [0, 0, 0, 0]

    frame_main = tkinter.Frame(self.window)
    frame_main.grid(column=0, row=0)

    frame_btn = tkinter.Frame(frame_main)
    frame_btn.grid(column=0, row=0)
    frame_picture = tkinter.Frame(frame_main)
    frame_picture.grid(column=1, row=0)

    frame_border_frame_btn_symbol = tkinter.Frame(frame_btn, bg="black")
    frame_border_frame_btn_symbol.grid(column=0, row=0, padx=5, )
    frame_border_frame_btn_point = tkinter.Frame(frame_btn, bg="black")
    frame_border_frame_btn_point.grid(column=0, row=1, padx=5, )
    frame_btn_symbol = tkinter.Frame(frame_border_frame_btn_symbol)
    frame_btn_symbol.grid(column=0, row=0, padx=3, pady=3)
    frame_btn_point = tkinter.Frame(frame_border_frame_btn_point)
    frame_btn_point.grid(column=0, row=0, padx=3, pady=3)
    frame_btn_other = tkinter.Frame(frame_btn)
    frame_btn_other.grid(column=0, row=2, padx=3, pady=3)

    label_btn_symbol = tkinter.Label(frame_btn_symbol, text="記号の位置指定")
    label_btn_symbol.grid(row=0, column=0, columnspan=3, sticky="we")   
    btn_set_symbol_nw = tkinter.Button(frame_btn_symbol, width=6, text="左上", command=functools.partial(set_position, "symbol", "position", "nw"))
    btn_set_symbol_nw.grid(column=0, row=1)
    btn_set_symbol_n  = tkinter.Button(frame_btn_symbol, width=6, text="上", command=functools.partial(set_position, "symbol", "position", "n"))
    btn_set_symbol_n.grid(column=1, row=1)
    btn_set_symbol_ne = tkinter.Button(frame_btn_symbol, width=6, text="右上", command=functools.partial(set_position, "symbol", "position", "ne"))
    btn_set_symbol_ne.grid(column=2, row=1)
    btn_set_symbol_w  = tkinter.Button(frame_btn_symbol, width=6, text="左", command=functools.partial(set_position, "symbol", "position", "w"))
    btn_set_symbol_w.grid(column=0, row=2)
    btn_set_symbol_c  = tkinter.Button(frame_btn_symbol, width=6, text="中央", command=functools.partial(set_position, "symbol", "position", "c"))
    btn_set_symbol_c.grid(column=1, row=2)
    btn_set_symbol_e  = tkinter.Button(frame_btn_symbol, width=6, text="右", command=functools.partial(set_position, "symbol", "position", "e"))
    btn_set_symbol_e.grid(column=2, row=2)
    btn_set_symbol_sw = tkinter.Button(frame_btn_symbol, width=6, text="左下", command=functools.partial(set_position, "symbol", "position", "sw"))
    btn_set_symbol_sw.grid(column=0, row=3)
    btn_set_symbol_s  = tkinter.Button(frame_btn_symbol, width=6, text="下", command=functools.partial(set_position, "symbol", "position", "s"))
    btn_set_symbol_s.grid(column=1, row=3)
    btn_set_symbol_se = tkinter.Button(frame_btn_symbol, width=6, text="右下", command=functools.partial(set_position, "symbol", "position", "se"))
    btn_set_symbol_se.grid(column=2, row=3)
    label_symbol_x = tkinter.Label(frame_btn_symbol, width=6, text="横位置")
    label_symbol_x.grid(column=0, row=4)
    position_symbol_x = tkinter.StringVar()
    position_symbol_x.set(dict_project["export"]["symbol"]["x"])
    entry_symbol_x = tkinter.Entry(frame_btn_symbol, width=5, justify=tkinter.CENTER)
    entry_symbol_x.delete(0, tkinter.END)
    entry_symbol_x.insert(0, position_symbol_x.get())
    entry_symbol_x.grid(column=1,  row=4, columnspan=2, sticky="we")
    label_symbol_move_y = tkinter.Label(frame_btn_symbol, width=6, text="縦位置")
    label_symbol_move_y.grid(column=0, row=5)
    position_symbol_y = tkinter.StringVar()
    position_symbol_y.set(dict_project["export"]["symbol"]["y"])
    entry_symbol_y = tkinter.Entry(frame_btn_symbol, width=5, justify=tkinter.CENTER)
    entry_symbol_y.delete(0, tkinter.END)
    entry_symbol_y.insert(0, position_symbol_y.get())
    entry_symbol_y.grid(column=1,  row=5, columnspan=2, sticky="we")
    label_symbol_size = tkinter.Label(frame_btn_symbol, width=6, text="大きさ")
    label_symbol_size.grid(column=0, row=6)
    position_symbol_size = tkinter.StringVar()
    position_symbol_size.set(dict_project["export"]["symbol"]["size"])
    entry_symbol_size = tkinter.Entry(frame_btn_symbol, width=5, justify=tkinter.CENTER)
    entry_symbol_size.delete(0, tkinter.END)
    entry_symbol_size.insert(0, position_symbol_size.get())
    entry_symbol_size.grid(column=1,  row=6, columnspan=2, sticky="we")
    booleanvar_unscored_symbol = tkinter.BooleanVar()
    booleanvar_correct_symbol = tkinter.BooleanVar()
    booleanvar_partial_symbol = tkinter.BooleanVar()
    booleanvar_hold_symbol = tkinter.BooleanVar()
    booleanvar_incorrect_symbol = tkinter.BooleanVar()
    chackbtn_unscored_symbol = tkinter.Checkbutton(frame_btn_symbol, text="未採点に未を表示", anchor=tkinter.W, variable=booleanvar_unscored_symbol, command=functools.partial(set_position, "symbol", "unscored", None))
    chackbtn_unscored_symbol.grid(column=0, row=7, columnspan=3, sticky="we", padx=3)
    chackbtn_correct_symbol = tkinter.Checkbutton(frame_btn_symbol, text="正答に○を表示", anchor=tkinter.W, variable=booleanvar_correct_symbol, command=functools.partial(set_position, "symbol", "correct", None))
    chackbtn_correct_symbol.grid(column=0, row=8, columnspan=3, sticky="we", padx=3)
    chackbtn_partial_symbol = tkinter.Checkbutton(frame_btn_symbol, text="部分点に△を表示", anchor=tkinter.W, variable=booleanvar_partial_symbol, command=functools.partial(set_position, "symbol", "partial", None))
    chackbtn_partial_symbol.grid(column=0, row=9, columnspan=3, sticky="we", padx=3)
    chackbtn_hold_symbol = tkinter.Checkbutton(frame_btn_symbol, text="保留に？を表示", anchor=tkinter.W, variable=booleanvar_hold_symbol, command=functools.partial(set_position, "symbol", "hold", None))
    chackbtn_hold_symbol.grid(column=0, row=10, columnspan=3, sticky="we", padx=3)
    chackbtn_incorrect_symbol = tkinter.Checkbutton(frame_btn_symbol, text="誤答に×を表示", anchor=tkinter.W, variable=booleanvar_incorrect_symbol, command=functools.partial(set_position, "symbol", "incorrect", None))
    chackbtn_incorrect_symbol.grid(column=0, row=11, columnspan=3, sticky="we", padx=3)

    label_btn_point = tkinter.Label(frame_btn_point, text="点数の位置指定")
    label_btn_point.grid(row=0, column=0, columnspan=3, sticky="we")   
    btn_set_point_nw = tkinter.Button(frame_btn_point, width=6, text="左上", command=functools.partial(set_position, "point", "position", "nw"))
    btn_set_point_nw.grid(column=0, row=1)
    btn_set_point_n  = tkinter.Button(frame_btn_point, width=6, text="上", command=functools.partial(set_position, "point", "position", "n"))
    btn_set_point_n.grid(column=1, row=1)
    btn_set_point_ne = tkinter.Button(frame_btn_point, width=6, text="右上", command=functools.partial(set_position, "point", "position", "ne"))
    btn_set_point_ne.grid(column=2, row=1)
    btn_set_point_w  = tkinter.Button(frame_btn_point, width=6, text="左", command=functools.partial(set_position, "point", "position", "w"))
    btn_set_point_w.grid(column=0, row=2)
    btn_set_point_c  = tkinter.Button(frame_btn_point, width=6, text="中央", command=functools.partial(set_position, "point", "position", "c"))
    btn_set_point_c.grid(column=1, row=2)
    btn_set_point_e  = tkinter.Button(frame_btn_point, width=6, text="右", command=functools.partial(set_position, "point", "position", "e"))
    btn_set_point_e.grid(column=2, row=2)
    btn_set_point_sw = tkinter.Button(frame_btn_point, width=6, text="左下", command=functools.partial(set_position, "point", "position", "sw"))
    btn_set_point_sw.grid(column=0, row=3)
    btn_set_point_s  = tkinter.Button(frame_btn_point, width=6, text="下", command=functools.partial(set_position, "point", "position", "s"))
    btn_set_point_s.grid(column=1, row=3)
    btn_set_point_se = tkinter.Button(frame_btn_point, width=6, text="右下", command=functools.partial(set_position, "point", "position", "se"))
    btn_set_point_se.grid(column=2, row=3)
    label_point_x = tkinter.Label(frame_btn_point, width=6, text="横位置")
    label_point_x.grid(column=0, row=4)
    position_point_x = tkinter.StringVar()
    position_point_x.set(dict_project["export"]["point"]["x"])
    entry_point_x = tkinter.Entry(frame_btn_point, width=5, justify=tkinter.CENTER, validate="key", validatecommand=(self.window.register(set_position), "point", "x", "%P"))
    entry_point_x.delete(0, tkinter.END)
    entry_point_x.insert(0, position_point_x.get())
    entry_point_x.grid(column=1,  row=4, columnspan=2, sticky="we")
    label_point_move_y = tkinter.Label(frame_btn_point, width=6, text="縦位置")
    label_point_move_y.grid(column=0, row=5)
    position_point_y = tkinter.StringVar()
    position_point_y.set(dict_project["export"]["point"]["y"])
    entry_point_y = tkinter.Entry(frame_btn_point, width=5, justify=tkinter.CENTER, validate="key", validatecommand=(self.window.register(set_position), "point", "y", "%P"))
    entry_point_y.delete(0, tkinter.END)
    entry_point_y.insert(0, position_point_y.get())
    entry_point_y.grid(column=1,  row=5, columnspan=2, sticky="we")
    label_point_size = tkinter.Label(frame_btn_point, width=6, text="大きさ")
    label_point_size.grid(column=0, row=6)
    position_point_size = tkinter.StringVar()
    position_point_size.set(dict_project["export"]["point"]["size"])
    entry_point_size = tkinter.Entry(frame_btn_point, width=5, justify=tkinter.CENTER, validate="key", validatecommand=(self.window.register(set_position), "point", "size", "%P"))
    entry_point_size.delete(0, tkinter.END)
    entry_point_size.insert(0, position_point_size.get())
    entry_point_size.grid(column=1,  row=6, columnspan=2, sticky="we")
    label_point_size_frame_btn = tkinter.Frame(frame_border_frame_btn_point)
    label_point_size_frame_btn.grid(column=1, row=6, columnspan=2, sticky="we")
    booleanvar_unscored_point = tkinter.BooleanVar()
    booleanvar_correct_point = tkinter.BooleanVar()
    booleanvar_partial_point = tkinter.BooleanVar()
    booleanvar_hold_point = tkinter.BooleanVar()
    booleanvar_incorrect_point = tkinter.BooleanVar()
    chackbtn_unscored_point = tkinter.Checkbutton(frame_btn_point, text="未採点に0を表示", anchor=tkinter.W, variable=booleanvar_unscored_point, command=functools.partial(set_position, "point", "unscored", None))
    chackbtn_unscored_point.grid(column=0, row=7, columnspan=3, sticky="we", padx=3)
    chackbtn_correct_point = tkinter.Checkbutton(frame_btn_point, text="正答に配点を表示", anchor=tkinter.W, variable=booleanvar_correct_point, command=functools.partial(set_position, "point", "correct", None))
    chackbtn_correct_point.grid(column=0, row=8, columnspan=3, sticky="we", padx=3)
    chackbtn_partial_point = tkinter.Checkbutton(frame_btn_point, text="部分点に点数を表示", anchor=tkinter.W, variable=booleanvar_partial_point, command=functools.partial(set_position, "point", "partial", None))
    chackbtn_partial_point.grid(column=0, row=9, columnspan=3, sticky="we", padx=3)
    chackbtn_hold_point = tkinter.Checkbutton(frame_btn_point, text="保留に点数を表示", anchor=tkinter.W, variable=booleanvar_hold_point, command=functools.partial(set_position, "point", "hold", None))
    chackbtn_hold_point.grid(column=0, row=10, columnspan=3, sticky="we", padx=3)
    chackbtn_incorrect_point = tkinter.Checkbutton(frame_btn_point, text="誤答に0を表示", anchor=tkinter.W, variable=booleanvar_incorrect_point, command=functools.partial(set_position, "point", "incorrect", None))
    chackbtn_incorrect_point.grid(column=0, row=11, columnspan=3, sticky="we", padx=3)

    btn_ex1 = tkinter.Button(frame_btn_other, width=6, text="例1", command=set_position_ex1)
    btn_ex1.grid(column=0, row=0)
    btn_ex2 = tkinter.Button(frame_btn_other, width=6, text="例2", command=set_position_ex2)
    btn_ex2.grid(column=1, row=0)
    btn_ex3 = tkinter.Button(frame_btn_other, width=6, text="例3", command=nothing_to_do)
    btn_ex3.grid(column=2, row=0)
    btn_export_picture = tkinter.Button(frame_btn_other, width=21, text="採点済答案画像の出力", bg="#ffbfbf", command=export_pdf)
    btn_export_picture.grid(column=0, row=1, columnspan=3)
    btn_export_xlsx = tkinter.Button(frame_btn_other, width=21, text="採点結果一覧表(.xlsx)の出力", bg="#bfffbf", command=export_list_xlsx)
    btn_export_xlsx.grid(column=0, row=2, columnspan=3)
    btn_help = tkinter.Button(frame_btn_other, width=21, text="ヘルプ", command=nothing_to_do)
    btn_help.grid(column=0, row=3, columnspan=3)
    btn_back = tkinter.Button(frame_btn_other, width=21, text="戻る", command=self.this_window_close)
    btn_back.grid(column=0, row=4, columnspan=3)

    frame_canvas = tkinter.Frame(frame_picture)
    frame_canvas.pack()

    canvas = tkinter.Canvas(frame_canvas, bg="black", width=567, height=800)
    canvas.bind("<Control-MouseWheel>", lambda eve:canvas.xview_scroll(int(-eve.delta/120), 'units'))
    canvas.bind("<MouseWheel>", lambda eve:canvas.yview_scroll(int(-eve.delta/120), 'units'))
    self.tk_image_model_answer = PIL.ImageTk.PhotoImage(file=path_file_model_answer)
    canvas.create_image(0, 0, image=self.tk_image_model_answer, anchor="nw")
    yscrollbar_canvas = tkinter.Scrollbar(frame_canvas, orient=tkinter.VERTICAL, command=canvas.yview)
    xscrollbar_canvas = tkinter.Scrollbar(frame_canvas, orient=tkinter.HORIZONTAL, command=canvas.xview)
    yscrollbar_canvas.pack(side="right", fill="y")
    xscrollbar_canvas.pack(side="bottom", fill="x")
    canvas.pack()
    canvas.config(
      xscrollcommand=xscrollbar_canvas.set,
      yscrollcommand=yscrollbar_canvas.set,
      scrollregion=(0, 0, self.tk_image_model_answer.width(), self.tk_image_model_answer.height())
    )
    preview_export_picture()
    


class MainFrame(tkinter.Frame):
  def __init__(self, root):
    super().__init__(root, width=800, height=500, borderwidth=2, relief="groove")
    self.root = root
    self.sub_window = SubWindow(self.root)
    self.pack()
    self.index_selected_exam = tkinter.IntVar(root)
    self.pack_propagate(0)
    self.create_listbox()
    self.btn_left()
    self.load_listbox_projects()

  # 試験一覧
  def create_listbox(self):
    frame_listbox = tkinter.Label(self)
    frame_listbox.grid(column=0, row=0, padx=5, pady=5)

    # 上部: ラベル
    label_listbox_header = tkinter.Label(frame_listbox, text="試験一覧", anchor="w")
    label_listbox_header.grid(column=0, row=0)

    # 中部: リストボックス
    self.listbox_projects = tkinter.Listbox(frame_listbox, width=60, height=20)
    self.listbox_projects.grid(column=0, row=1)
    self.listbox_projects.configure(
      activestyle=tkinter.DOTBOX,
      selectmode=tkinter.SINGLE,
      selectbackground="grey"
    )
    self.listbox_projects.bind("<<ListboxSelect>>", self.selected_element_in_listbox)

    # 下部: ボタン
    frame_listbox_footer = tkinter.Frame(frame_listbox)
    frame_listbox_footer.grid(column=0, row=2)
    tkinter.Button(frame_listbox_footer, text="追加", width=10, height=1, command=self.sub_window.add_project).grid(column=0, row=0)
    tkinter.Button(frame_listbox_footer, text="編集", width=10, height=1, command=self.sub_window.edit_project).grid(column=1, row=0)
    tkinter.Button(frame_listbox_footer, text="削除", width=10, height=1, command=self.del_project).grid(column=2, row=0)
    tkinter.Button(frame_listbox_footer, text="上へ", width=10, height=1, command=self.up_project).grid(column=3, row=0)
    tkinter.Button(frame_listbox_footer, text="下へ", width=10, height=1, command=self.down_project).grid(column=4, row=0)

  def write_index_to_config(self, index_projects_in_listbox):
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    dict_config["index_projects_in_listbox"] = index_projects_in_listbox
    with open("config.json", "w", encoding="utf-8") as f:
      json.dump(dict_config, f, indent=2)

  def selected_element_in_listbox(self, event):
    if self.listbox_projects.curselection() != ():
      index_projects_in_listbox = self.listbox_projects.curselection()[0]
      self.write_index_to_config(index_projects_in_listbox)

  def load_listbox_projects(self, *, parent=None):
    if parent is not None:
      self = parent
    self.listbox_projects.delete(0, tkinter.END)
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    if len(dict_config["projects"]) == 0:
      self.listbox_projects.insert(0, "［追加］をクリックして新しく試験を追加して下さい")
      self.listbox_projects.configure(state="disable")
      self.write_index_to_config(None)
    else:
      for project in dict_config["projects"]:
        self.listbox_projects.insert(tkinter.END, project["name"])
      index_projects_in_listbox = dict_config["index_projects_in_listbox"]
      self.listbox_projects.select_set(index_projects_in_listbox)
  
  def del_project(self):
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    index_projects_in_listbox = dict_config["index_projects_in_listbox"]
    if index_projects_in_listbox is None:
      tkinter.messagebox.showinfo(
        "試験が選択されていません. ", 
        "「試験一覧」より削除したい試験を選択して下さい. "
      )
    else:
      bool_del_project = tkinter.messagebox.askyesno(
        "試験を削除しますか？",
        f"この操作で, 答案スキャンデータ / 採点データが失われることはありませんが, 試験一覧からは表示されなくなり, 本アプリ上からはアクセスできなくなります. \n"
        + f"［追加］より同じフォルダ / ファイルを指定することで, 採点データ等を再び利用できます. \n"
        + f"採点データ等を完全に削除したい場合は, 本アプリ終了後, 答案スキャンデータが保存されているフォルダ内にある隠しフォルダ「.temp_saiten」を手動で削除して下さい. \n\n"
        + f"試験名: {dict_config['projects'][index_projects_in_listbox]['name']}\n\n"
        + f"本当に試験を削除しますか？"
      )
      if bool_del_project:
        with open("config.json", "r", encoding="utf-8") as f:
          dict_config = json.load(f)
        dict_config["projects"].pop(index_projects_in_listbox)
        if len(dict_config["projects"]) == 0:
          dict_config["index_projects_in_listbox"] = None
        else:
          dict_config["index_projects_in_listbox"] = 0
        with open("config.json", "w", encoding="utf-8") as f:
          json.dump(dict_config, f, indent=2)
        self.load_listbox_projects()
  
  def up_project(self):
    nothing_to_do()
    self.load_listbox_projects()

  def down_project(self):
    nothing_to_do()
    self.load_listbox_projects()

  def make_xlsx(self):
    tkinter.messagebox.showinfo(
      "配点を入力します",
      "配点の入力は, 本ソフトウェア上ではなく Excel 等の表計算ソフトウェアを使用して行います. \n\n"
      + "配点を登録するために 名簿と配点の入力.xlsx ファイルを作成して開きます. \n\n"
      + "作成には数十秒かかる場合があります. \n"
      + "自動的に Excel が起動するまで操作しないで下さい. "
    )
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
    path_dir = dict_project["path_dir"]
    with open(path_dir + "/.temp_saiten/answer_area.json") as f:
      dict_answer_area = json.load(f)
    with open(path_dir + "/.temp_saiten/load_picture.json") as f:
      dict_load_picture = json.load(f)
      if os.path.exists(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx"):
        bool_del_xlsx = tkinter.messagebox.askokcancel(
          "配点ファイルが存在しています", 
          "配点ファイルに入力した情報を保存するには, ［配点を読み込む］をクリックする必要があります. \n"
          + "既に Excel で配点を入力されている場合で［配点を読み込む］をクリックしていない場合は, 入力した情報が破棄されます. \n\n"
          + "入力した配点を保存した上で操作を続行したい場合は, ［キャンセル］をクリックした後, ［配点を読み込む］をクリックして配点を読み込んでから, もう一度実行して下さい. \n\n"
          + "配点ファイルを削除してもよろしいですか？"
        )
        if not bool_del_xlsx:
          return None

    if len(dict_answer_area["questions"]) == 0:
      tkinter.messagebox.showwarning(
        "解答欄の位置が指定されていません",
        "解答欄の位置が指定されていません. \n"
        + "［解答欄の位置を指定］をクリックして解答欄の位置を指定してから, もう一度お試し下さい. "
      )
      return
    if os.path.exists(path_dir + "/.temp_saiten/meibo.json"):
      with open(path_dir + "/.temp_saiten/meibo.json", "r", encoding="utf-8") as f:
        list_meibo = json.load(f)
    else:
      list_meibo = []
    for i in range(len(dict_load_picture["answer"]) - len(list_meibo)):
      list_meibo.append(
        {
          "学年": "",
          "学級": "",
          "出席番号": "",
          "生徒番号": "",
          "氏名": ""
        }
      )
    with open(path_dir + "/.temp_saiten/meibo.json", "w", encoding="utf-8") as f:
      json.dump(list_meibo, f, indent=2)
    
    workbook_import = openpyxl.Workbook()
    workbook_import.remove(workbook_import["Sheet"])
    workbook_import.create_sheet(title="名簿登録")
    sheet_meibo = workbook_import["名簿登録"]
    sheet_meibo.freeze_panes = ("B2")
    sheet_meibo.cell(1, 1).value = "答案番号"
    sheet_meibo.row_dimensions[1].height = 40
    for index_row in range(len(list_meibo)):
      sheet_meibo.cell(index_row + 2, 1).value = index_row
      sheet_meibo.row_dimensions[index_row + 2].height = 30
    for index_column, key in enumerate(["学年", "学級", "出席番号", "生徒番号", "氏名"]):
      sheet_meibo.cell(1, index_column + 2).value = key
      if key in ["学年", "学級", "出席番号"]:
        sheet_meibo.column_dimensions[openpyxl.utils.cell.get_column_letter(index_column + 2)].width = 10
      else:
        sheet_meibo.column_dimensions[openpyxl.utils.cell.get_column_letter(index_column + 2)].width = 20
      for index_row in range(len(list_meibo)):
        sheet_meibo.cell(index_row + 2, index_column + 2).value = list_meibo[index_row][key]
        if index_column + 2 in [2, 4]:
          color_background = "bfffff"
        elif index_column + 2 in [3]:
          color_background = "cccccc"
        elif index_column + 2 in [5]:
          color_background = "ffbfbf"
        elif index_column + 2 in [6]:
          color_background = "ffdfdf"          
        sheet_meibo.cell(index_row + 2, index_column + 2).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor=color_background)
        sheet_meibo.cell(index_row + 2, index_column + 2).protection = openpyxl.styles.Protection(locked=False)
    list_add_images = []
    index_column += 2
    for str_type in ["生徒番号", "氏名"]:
      for question in dict_answer_area["questions"]:
        if question["type"] == str_type:
          index_column += 1
          list_add_images.append([])
          sheet_meibo.cell(1, index_column).value = f"({str_type})"
          for index_meibo, meibo in enumerate(list_meibo):
            list_add_images[-1].append(PIL.Image.open(f"{path_dir}/.temp_saiten/answer/{index_meibo}.png"))
            list_add_images[-1][-1] = list_add_images[-1][-1].crop((question["area"][0], question["area"][1], question["area"][2], question["area"][3]))
            height_image = 40
            width_image = list_add_images[-1][-1].width * 40 // list_add_images[-1][-1].height
            list_add_images[-1][-1] = list_add_images[-1][-1].resize((width_image, height_image))
            list_add_images[-1][-1].save(f"{path_dir}/.temp_saiten/make_xlsx/{index_column}_{index_meibo}.png")
            list_add_images[-1][-1] = openpyxl.drawing.image.Image(f"{path_dir}/.temp_saiten/make_xlsx/{index_column}_{index_meibo}.png")
            sheet_meibo.add_image(list_add_images[-1][-1], f"{openpyxl.utils.get_column_letter(index_column)}{index_meibo + 2}")
          sheet_meibo.column_dimensions[openpyxl.utils.get_column_letter(index_column)].width = width_image / 8
    workbook_import.create_sheet(title="配点登録")
    sheet_haiten = workbook_import["配点登録"]
    sheet_haiten.cell(1, 1).value = "枠番号"
    sheet_haiten.cell(1, 2).value = "種類"
    sheet_haiten.cell(1, 3).value = "大問"
    sheet_haiten.cell(1, 4).value = "小問"
    sheet_haiten.cell(1, 5).value = "枝問"
    sheet_haiten.cell(1, 6).value = "配点"
    side = openpyxl.styles.Side(style="thin", color="000000")
    border_up_down = openpyxl.styles.Border(top=side, bottom=side)
    datavalidation_whole = openpyxl.worksheet.datavalidation.DataValidation(type="whole")
    datavalidation_textlength10 = openpyxl.worksheet.datavalidation.DataValidation(type="textLength", operator="lessThanOrEqual", formula1=10)
    
    sheet_haiten.row_dimensions[1].height = 22.5
    for index_question, question in enumerate(dict_answer_area["questions"]):
      sheet_haiten.row_dimensions[index_question + 2].height = 22.5
      sheet_haiten.cell(index_question + 2, 1).value = index_question
      sheet_haiten.cell(index_question + 2, 1).border = border_up_down
      sheet_haiten.cell(index_question + 2, 2).value = question["type"]
      sheet_haiten.cell(index_question + 2, 2).border = border_up_down
      sheet_haiten.cell(index_question + 2, 3).value = question["daimon"]
      sheet_haiten.cell(index_question + 2, 3).border = border_up_down
      if question["type"] in ["設問", "小計点"]:
        sheet_haiten.cell(index_question + 2, 3).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="bfffff")
        sheet_haiten.cell(index_question + 2, 3).protection = openpyxl.styles.Protection(locked=False)
        datavalidation_textlength10.add(sheet_haiten.cell(index_question + 2, 3))
      else:
        sheet_haiten.cell(index_question + 2, 3).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="cccccc")
      sheet_haiten.cell(index_question + 2, 4).value = question["shomon"]
      sheet_haiten.cell(index_question + 2, 4).border = border_up_down
      if question["type"] in ["設問"]:
        sheet_haiten.cell(index_question + 2, 4).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="cfefef")
        sheet_haiten.cell(index_question + 2, 4).protection = openpyxl.styles.Protection(locked=False)
        datavalidation_textlength10.add(sheet_haiten.cell(index_question + 2, 4))
      else:
        sheet_haiten.cell(index_question + 2, 4).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="cccccc")
      sheet_haiten.cell(index_question + 2, 5).value = question["shimon"]
      sheet_haiten.cell(index_question + 2, 5).border = border_up_down
      if question["type"] in ["設問"]:
        sheet_haiten.cell(index_question + 2, 5).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="bfffff")
        sheet_haiten.cell(index_question + 2, 5).protection = openpyxl.styles.Protection(locked=False)
        datavalidation_textlength10.add(sheet_haiten.cell(index_question + 2, 5))
      else:
        sheet_haiten.cell(index_question + 2, 5).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="cccccc")
      sheet_haiten.cell(index_question + 2, 6).value = question["haiten"]
      sheet_haiten.cell(index_question + 2, 6).border = border_up_down
      if question["type"] in ["設問"]:
        sheet_haiten.cell(index_question + 2, 6).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="ffbfbf")
        sheet_haiten.cell(index_question + 2, 6).protection = openpyxl.styles.Protection(locked=False)
        datavalidation_whole.add(sheet_haiten.cell(index_question + 2, 6))
      else:
        sheet_haiten.cell(index_question + 2, 6).fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="cccccc")
    sheet_haiten.cell(index_question + 3, 5).value = "配点合計"
    sheet_haiten.cell(index_question + 3, 6).value = f"=SUMIF(B2:B{index_question + 2}, \"設問\", F2:F{index_question + 2})"
    for sheet in workbook_import.worksheets:
      for row in sheet.rows:
        for cell in row:
          cell.font = openpyxl.styles.fonts.Font(size=11, name="Meiryo UI")
          cell.alignment = openpyxl.styles.alignment.Alignment(horizontal="center", vertical="center")

    sheet_meibo.protection.selectLockedCells   = True  # ロックされたセルの選択
    sheet_meibo.protection.selectUnlockedCells = False # ロックされていないセルの選択
    sheet_meibo.protection.formatCells         = True  # セルの書式設定
    sheet_meibo.protection.formatColumns       = True  # 列の書式設定
    sheet_meibo.protection.formatRows          = True  # 行の書式設定
    sheet_meibo.protection.insertColumns       = True  # 列の挿入
    sheet_meibo.protection.insertRows          = True  # 行の挿入
    sheet_meibo.protection.insertHyperlinks    = True  # ハイパーリンクの挿入
    sheet_meibo.protection.deleteColumns       = True  # 列の削除
    sheet_meibo.protection.deleteRows          = True  # 行の削除
    sheet_meibo.protection.sort                = True  # 並べ替え
    sheet_meibo.protection.autoFilter          = True  # フィルター
    sheet_meibo.protection.pivotTables         = True  # ピボットテーブルレポート
    sheet_meibo.protection.objects             = True  # オブジェクトの編集
    sheet_meibo.protection.scenarios           = True  # シナリオの編集
    sheet_meibo.protection.enable()
    sheet_haiten.protection.selectLockedCells   = True  # ロックされたセルの選択
    sheet_haiten.protection.selectUnlockedCells = False # ロックされていないセルの選択
    sheet_haiten.protection.formatCells         = True  # セルの書式設定
    sheet_haiten.protection.formatColumns       = True  # 列の書式設定
    sheet_haiten.protection.formatRows          = True  # 行の書式設定
    sheet_haiten.protection.insertColumns       = True  # 列の挿入
    sheet_haiten.protection.insertRows          = True  # 行の挿入
    sheet_haiten.protection.insertHyperlinks    = True  # ハイパーリンクの挿入
    sheet_haiten.protection.deleteColumns       = True  # 列の削除
    sheet_haiten.protection.deleteRows          = True  # 行の削除
    sheet_haiten.protection.sort                = True  # 並べ替え
    sheet_haiten.protection.autoFilter          = True  # フィルター
    sheet_haiten.protection.pivotTables         = True  # ピボットテーブルレポート
    sheet_haiten.protection.objects             = True  # オブジェクトの編集
    sheet_haiten.protection.scenarios           = True  # シナリオの編集
    sheet_haiten.protection.enable()
    workbook_import.security.lockStructure = True

    try:
      workbook_import.save(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx")
    except PermissionError:
      tkinter.messagebox.showerror(
        "ファイルを保存できません",
        "ファイルを保存できませんでした. \n"
        + "既にファイルを開いていませんか？\n"
        + "Excel を終了して, もう一度お試し下さい. "
      )
    else:
      os.startfile(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx")

  def read_xlsx(self):
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    dict_project = dict_config["projects"][dict_config["index_projects_in_listbox"]]
    path_dir = dict_project["path_dir"]
    with open(path_dir + "/.temp_saiten/answer_area.json") as f:
      dict_answer_area = json.load(f)
    with open(path_dir + "/.temp_saiten/meibo.json", "r", encoding="utf-8") as f:
      list_meibo = json.load(f)
    if not os.path.exists(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx"):
      tkinter.messagebox.showerror(
        "ファイルが見つかりません",
        "名簿と配点の入力.xlsx が見つかりません. \n"
        + "［配点を入力する］をクリックして, ファイルを生成し, 配点を入力して保存して下さい. "
      )
    else:
      try:
        workbook_import = openpyxl.load_workbook(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx", data_only=True)
        os.remove(path_dir + "/.temp_saiten/名簿と配点の入力.xlsx")
      except PermissionError:
        tkinter.messagebox.showerror(
          "ファイルを操作できません",
          "ファイルを操作できませんでした. \n"
          + "ファイルを開いていませんか？\n"
          + "Excel を終了して, もう一度お試し下さい. "
        )
        return
      sheet_meibo = workbook_import["名簿登録"]
      for index_meibo in range(len(list_meibo)):
        list_meibo[index_meibo]["学年"] = sheet_meibo.cell(index_meibo + 2, 2).value
        list_meibo[index_meibo]["学級"] = sheet_meibo.cell(index_meibo + 2, 3).value
        list_meibo[index_meibo]["出席番号"] = sheet_meibo.cell(index_meibo + 2, 4).value
        list_meibo[index_meibo]["生徒番号"] = sheet_meibo.cell(index_meibo + 2, 5).value
        list_meibo[index_meibo]["氏名"] = sheet_meibo.cell(index_meibo + 2, 6).value
      sheet_haiten = workbook_import["配点登録"]
      for index_question in range(len(dict_answer_area["questions"])):
        dict_answer_area["questions"][index_question]["daimon"] = sheet_haiten.cell(index_question + 2, 3).value
        dict_answer_area["questions"][index_question]["shomon"] = sheet_haiten.cell(index_question + 2, 4).value
        dict_answer_area["questions"][index_question]["shimon"] = sheet_haiten.cell(index_question + 2, 5).value
        if sheet_haiten.cell(index_question + 2, 5).value == "":
          dict_answer_area["questions"][index_question]["haiten"] = None
        else:
          dict_answer_area["questions"][index_question]["haiten"] = sheet_haiten.cell(index_question + 2, 6).value
      with open(path_dir + "/.temp_saiten/meibo.json", "w", encoding="utf-8") as f:
        json.dump(list_meibo, f, indent=2)
      with open(path_dir + "/.temp_saiten/answer_area.json", "w", encoding="utf-8") as f:
        json.dump(dict_answer_area, f, indent=2)
      tkinter.messagebox.showinfo(
        "配点を読み込みました",
        "読み込んだ内容は保存し, 名簿と配点の入力.xlsx は削除しました. \n"
        + "再び配点を編集するには, ［配点を入力する］をクリックして下さい. \n"
      )

  # btn_left: 操作ボタン
  def btn_left(self):
    frame_operate = tkinter.Frame(self)
    frame_operate.grid(column=1, row=0, padx=10, pady=10)
    tkinter.Button(frame_operate, text="解答欄の位置を指定", command=self.sub_window.select_area, width=20, height=2).grid(column=0, row=0, columnspan=2)
    tkinter.Button(frame_operate, text="名簿/配点を\nExcel で入力", command=self.make_xlsx, width=9, height=2).grid(column=0, row=1, sticky="WE")
    tkinter.Button(frame_operate, text="名簿/配点を\n読み込む", command=self.read_xlsx, width=9, height=2).grid(column=1, row=1, sticky="WE")
    tkinter.Frame(frame_operate, width=20, height=25).grid(column=0, row=2, columnspan=2)
    tkinter.Button(frame_operate, text="一括採点する", command=self.sub_window.score_answer, width=20, height=2).grid(column=0, row=3, columnspan=2)
    tkinter.Frame(frame_operate, width=20, height=25).grid(column=0, row=4, columnspan=2)
    tkinter.Button(frame_operate, text="書き出す", command=self.sub_window.export, width=20, height=2).grid(column=0, row=5, columnspan=2)
    tkinter.Button(frame_operate, text="終了", command=self.root.destroy, width=20, height=2).grid(column=0, row=6, columnspan=2)  

def menu(root):
  menu_root = tkinter.Menu(root)

  menu_file = tkinter.Menu(menu_root, tearoff=0)
  menu_file.add_command(label="新しく試験を追加")
  menu_file.add_command(label="選択中の試験を編集")
  menu_file.add_command(label="選択中の試験を削除")
  menu_file.add_separator()
  menu_file.add_command(label="構成設定をリセット")
  menu_file.add_separator()
  menu_file.add_command(label="終了")

  menu_edit = tkinter.Menu(menu_root, tearoff=0)
  menu_edit.add_command(label="選択中の試験の解答欄の位置を指定")
  menu_edit.add_command(label="選択中の試験の配点を入力する")
  menu_edit.add_command(label="選択中の試験の配点を読み込む")
  menu_edit.add_command(label="選択中の試験を一括採点する")

  menu_help = tkinter.Menu(menu_root, tearoff=0)
  menu_help.add_command(label="ヘルプ")
  menu_help.add_command(label="バージョン情報")
  
  menu_root.add_cascade(label="ファイル", menu=menu_file)
  menu_root.add_cascade(label="編集", menu=menu_edit)
  menu_root.add_cascade(label="ヘルプ", menu=menu_help)
  root.config(menu=menu_root)

def make_config():
  dict_config = {
    "index_projects_in_listbox": None,
    "projects": []
  }
  with open("config.json", "w", encoding="utf-8") as f:
    json.dump(dict_config, f, indent=2)

def check_on_run():
  try:
    with open("config.json", "r", encoding="utf-8") as f:
      dict_config = json.load(f)
    return True
  except FileNotFoundError:
    tkinter.messagebox.showinfo(
      "ごめんなさい",
      "本ソフトウェアは, alpha版です. \n\n"
      + "一部動作しない機能がございます. "
    )
    bool_accept_terms = tkinter.messagebox.askyesno(
      "Accept the terms? - 規約に同意しますか？",
      "Copyright(c) 2022 KeppyNaushika\n\n"
      + "This software is released under the GNU Affero General Public License v3.0, see LICENSE. \n\n"
      + "このソフトウェアは, GNU Affero General Public License version3 の下, 提供されています. \n\n"
      + "ライセンスを遵守する限り, 商用利用, 変更, 頒布, 特許利用, 私的利用が認められますが, "
      + "利用にあたって開発者は責任を負いませんしいかなる保証も提供しません. \n\n"
      + "同梱されている LICENSE をお読みいただき, 同意される場合は［はい］をクリックして下さい. \n\n"
      + "尚, 本ソフトウェアにおける Microsoft 製品についての記述は, マイクロソフトの商標およびブランドガイドラインに準拠しています. \n\n"
      + "The github repository of this software:\n"
      + "https://github.com/KeppyNaushika/scoring_at_once/"
    )
    if bool_accept_terms:
      tkinter.messagebox.showinfo(
        "表計算ソフトをご用意下さい. ",
        "本ソフトウェアでは, 一部で Microsoft Excel 等の表計算ソフトウェアを利用します. \n\n"
        + "あらかじめインストールの上, ご利用下さい. "
      )
      make_config()
      return True
    else:
      return False

def main():
  root = tkinter.Tk()
  root.title("一括採点 - alpha版")
  root.geometry("800x500")
  menu(root)
  MainFrame(root=root)
  root.mainloop()

if __name__ == "__main__":
  if check_on_run():
    main()